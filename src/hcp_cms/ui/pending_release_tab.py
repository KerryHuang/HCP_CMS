"""待發清單 Tab — 顯示待發布的 Patch 確認項目，支援月份篩選、異動月份與標記已發布。"""

from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path

from PySide6.QtCore import Qt, QUrl
from PySide6.QtGui import QDesktopServices
from PySide6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from hcp_cms.core.release_manager import ReleaseManager

# 欄位定義：(標題, ReleaseItem 屬性名稱, 是否 Stretch)
# 順序：狀態 → 客戶 → Mantis票號 → 修改者 → 指派人 → 備注 → 案件編號
_COLUMNS = [
    ("狀態",        "status",           False),
    ("客戶",        "client_name",      False),
    ("Mantis 票號", "mantis_ticket_id", False),
    ("修改者",      "modifier",         False),
    ("指派人",      "assignee",         False),
    ("備注",        "note",             True),
    ("案件編號",    "case_id",          False),
]

# 匯出 Excel 時的欄位順序（與 _COLUMNS 一致，方便維護）
_EXPORT_HEADERS = [c[0] for c in _COLUMNS]


class PendingReleaseTab(QWidget):
    """待發清單分頁：依月份顯示待發項目，可標記已發布或移至其他月份。"""

    def __init__(self, conn: sqlite3.Connection | None = None, parent=None) -> None:
        super().__init__(parent)
        self._conn = conn
        self._items: list = []
        self._setup_ui()
        if conn:
            self.refresh()

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(6)

        # ── 篩選列 ────────────────────────────────────────────────
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("檢視月份："))
        self._month_combo = QComboBox()
        self._month_combo.setMinimumWidth(120)
        self._populate_months(self._month_combo)
        self._month_combo.currentIndexChanged.connect(self.refresh)
        filter_row.addWidget(self._month_combo)

        refresh_btn = QPushButton("🔄 重新整理")
        refresh_btn.clicked.connect(self.refresh)
        filter_row.addWidget(refresh_btn)

        filter_row.addStretch()

        export_btn = QPushButton("📥 匯出 Excel")
        export_btn.setToolTip("將目前月份的待發清單匯出為 .xlsx 檔案")
        export_btn.clicked.connect(self._on_export_excel)
        filter_row.addWidget(export_btn)

        layout.addLayout(filter_row)

        # ── 操作列 ────────────────────────────────────────────────
        action_row = QHBoxLayout()

        self._up_btn = QPushButton("⬆ 上移")
        self._up_btn.setEnabled(False)
        self._up_btn.setToolTip("將選取項目往上移一列")
        self._up_btn.clicked.connect(self._on_move_up)
        action_row.addWidget(self._up_btn)

        self._down_btn = QPushButton("⬇ 下移")
        self._down_btn.setEnabled(False)
        self._down_btn.setToolTip("將選取項目往下移一列")
        self._down_btn.clicked.connect(self._on_move_down)
        action_row.addWidget(self._down_btn)

        action_row.addSpacing(12)

        self._confirm_btn = QPushButton("🔍 待確認")
        self._confirm_btn.setEnabled(False)
        self._confirm_btn.setToolTip("標記為待確認（介於待發與已發布之間）")
        self._confirm_btn.clicked.connect(self._on_mark_pending_confirm)
        action_row.addWidget(self._confirm_btn)

        self._release_btn = QPushButton("✅ 標記已發布")
        self._release_btn.setEnabled(False)
        self._release_btn.clicked.connect(self._on_mark_released)
        action_row.addWidget(self._release_btn)

        self._undo_btn = QPushButton("↩ 退回待發")
        self._undo_btn.setEnabled(False)
        self._undo_btn.clicked.connect(self._on_mark_pending)
        action_row.addWidget(self._undo_btn)

        self._delete_btn = QPushButton("🗑 刪除")
        self._delete_btn.setEnabled(False)
        self._delete_btn.clicked.connect(self._on_delete_item)
        action_row.addWidget(self._delete_btn)

        action_row.addSpacing(20)
        action_row.addWidget(QLabel("移至月份："))
        self._move_month_combo = QComboBox()
        self._move_month_combo.setMinimumWidth(120)
        self._populate_months(self._move_month_combo)
        action_row.addWidget(self._move_month_combo)

        self._move_btn = QPushButton("📅 確認移動")
        self._move_btn.setEnabled(False)
        self._move_btn.clicked.connect(self._on_move_month)
        action_row.addWidget(self._move_btn)

        action_row.addStretch()
        layout.addLayout(action_row)

        # ── 表格 ──────────────────────────────────────────────────
        self._table = QTableWidget(0, len(_COLUMNS))
        self._table.setHorizontalHeaderLabels([c[0] for c in _COLUMNS])
        hdr = self._table.horizontalHeader()
        for i, (_, _, stretch) in enumerate(_COLUMNS):
            if stretch:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        # 允許使用者拖拉調整欄位順序
        hdr.setSectionsMovable(True)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSortingEnabled(True)
        self._table.setWordWrap(True)
        self._table.itemSelectionChanged.connect(self._on_selection_changed)
        self._table.cellDoubleClicked.connect(self._on_cell_double_clicked)
        layout.addWidget(self._table)

    def _populate_months(self, combo: QComboBox) -> None:
        """填入未來 3 個月 + 當月 + 過去 12 個月，共 16 個月。"""
        now = datetime.now()
        combo.blockSignals(True)
        combo.clear()
        for i in range(-3, 13):
            m = now.month - i
            y = now.year
            while m <= 0:
                m += 12
                y -= 1
            while m > 12:
                m -= 12
                y += 1
            ms = f"{y}{m:02d}"
            combo.addItem(f"{ms[:4]}/{ms[4:]}", ms)
        combo.blockSignals(False)

    def refresh(self) -> None:
        if not self._conn:
            return
        month_str = self._month_combo.currentData()
        if not month_str:
            return
        mgr = ReleaseManager(self._conn)
        # 排序：待發(0) → 待確認(1) → 已發布(2)；各群組內保留手動 sort_order
        _STATUS_ORDER = {"待發": 0, "待確認": 1, "已發布": 2}
        raw = mgr.list_by_month(month_str)
        self._items = sorted(raw, key=lambda x: (
            _STATUS_ORDER.get(x.status, 9),
            x.sort_order if x.sort_order is not None else 999999,
        ))

        self._table.setSortingEnabled(False)
        self._table.setRowCount(0)
        for item in self._items:
            row = self._table.rowCount()
            self._table.insertRow(row)
            for col, (_, attr, _) in enumerate(_COLUMNS):
                val = getattr(item, attr) or ""
                cell = QTableWidgetItem(val)
                cell.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                if item.status == "已發布":
                    cell.setForeground(Qt.GlobalColor.gray)
                elif item.status == "待確認":
                    from PySide6.QtGui import QColor
                    cell.setForeground(QColor("#F59E0B"))  # 琥珀色
                self._table.setItem(row, col, cell)
        self._table.setSortingEnabled(True)
        self._table.resizeRowsToContents()
        self._release_btn.setEnabled(False)
        self._move_btn.setEnabled(False)

    # ── 匯出 Excel ────────────────────────────────────────────────

    def _on_export_excel(self) -> None:
        """將目前月份的清單匯出為 xlsx 檔案。"""
        if not self._items:
            QMessageBox.information(self, "提示", "目前月份無資料可匯出。")
            return

        month_str = self._month_combo.currentData() or "unknown"
        default_name = f"待發清單_{month_str}.xlsx"

        path, _ = QFileDialog.getSaveFileName(
            self,
            "匯出待發清單",
            str(Path.home() / default_name),
            "Excel 檔案 (*.xlsx)",
        )
        if not path:
            return

        try:
            self._export_to_xlsx(path, month_str)
            QMessageBox.information(self, "完成", f"已匯出至：\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "匯出失敗", str(e))

    def _export_to_xlsx(self, path: str, month_str: str) -> None:
        """實際寫入 xlsx 檔案。"""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{month_str[:4]}-{month_str[4:]} 待發清單"

        # ── 標題列 ────────────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx, header in enumerate(_EXPORT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── 資料列 ────────────────────────────────────────────────
        gray_font = Font(color="888888")
        amber_font = Font(color="B45309")   # 待確認：深琥珀色（避免淺色難讀）
        for row_idx, item in enumerate(self._items, start=2):
            for col_idx, (_, attr, _) in enumerate(_COLUMNS, start=1):
                val = getattr(item, attr) or ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if item.status == "已發布":
                    cell.font = gray_font
                elif item.status == "待確認":
                    cell.font = amber_font

        # ── 欄寬設定 ──────────────────────────────────────────────
        col_widths = {
            "狀態": 8,
            "客戶": 20,
            "Mantis 票號": 14,
            "修改者": 12,
            "指派人": 12,
            "備注": 50,
            "案件編號": 16,
        }
        for col_idx, header in enumerate(_EXPORT_HEADERS, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = col_widths.get(header, 15)

        # 凍結標題列
        ws.freeze_panes = "A2"

        wb.save(path)

    # ── 選取與操作 ────────────────────────────────────────────────

    def _on_selection_changed(self) -> None:
        rows = self._table.selectionModel().selectedRows()
        if not rows:
            self._up_btn.setEnabled(False)
            self._down_btn.setEnabled(False)
            self._confirm_btn.setEnabled(False)
            self._release_btn.setEnabled(False)
            self._undo_btn.setEnabled(False)
            self._delete_btn.setEnabled(False)
            self._move_btn.setEnabled(False)
            return
        row = rows[0].row()
        if 0 <= row < len(self._items):
            status = self._items[row].status
            self._up_btn.setEnabled(row > 0)
            self._down_btn.setEnabled(row < len(self._items) - 1)
            # 待發 → 可標記待確認或直接已發布
            # 待確認 → 可標記已發布或退回待發
            # 已發布 → 只能退回待發
            self._confirm_btn.setEnabled(status == "待發")
            self._release_btn.setEnabled(status in ("待發", "待確認"))
            self._undo_btn.setEnabled(status in ("待確認", "已發布"))
            self._delete_btn.setEnabled(True)
            self._move_btn.setEnabled(True)

    def _selected_item(self):
        rows = self._table.selectionModel().selectedRows()
        if not rows:
            return None
        row = rows[0].row()
        if 0 <= row < len(self._items):
            return self._items[row]
        return None

    def _on_move_up(self) -> None:
        self._shift_selected_item(-1)

    def _on_move_down(self) -> None:
        self._shift_selected_item(1)

    def _shift_selected_item(self, direction: int) -> None:
        """將選取列上移（-1）或下移（+1）。"""
        rows = self._table.selectionModel().selectedRows()
        if not rows or not self._conn:
            return
        row = rows[0].row()
        if row < 0 or row >= len(self._items):
            return
        ReleaseManager(self._conn).move_item(self._items, row, direction)
        self.refresh()
        # 重新選取移動後的位置
        new_row = row + direction
        if 0 <= new_row < self._table.rowCount():
            self._table.selectRow(new_row)

    def _on_cell_double_clicked(self, row: int, col: int) -> None:
        """雙擊不同欄位執行對應動作：
        - Mantis 票號欄：開啟 Mantis 網頁
        - 備注欄：彈出編輯視窗
        """
        if row < 0 or row >= len(self._items):
            return
        # 因 header 可移動，用 logicalIndex 對應欄位屬性
        logical_col = self._table.horizontalHeader().logicalIndex(col)
        col_attr = _COLUMNS[logical_col][1] if 0 <= logical_col < len(_COLUMNS) else ""

        item = self._items[row]

        if col_attr == "mantis_ticket_id":
            self._open_mantis_url(item.mantis_ticket_id)

        elif col_attr == "note":
            current_note = item.note or ""
            new_note, ok = QInputDialog.getMultiLineText(
                self, "編輯備注", f"#{item.mantis_ticket_id or item.case_id} 備注：", current_note
            )
            if ok and new_note != current_note and self._conn:
                ReleaseManager(self._conn).update_note(item.id, new_note)
                self.refresh()

    def _open_mantis_url(self, ticket_id: str | None) -> None:
        """依系統設定的 Mantis URL 開啟對應票單頁面。"""
        if not ticket_id:
            return
        try:
            from urllib.parse import urlparse

            from hcp_cms.services.credential import CredentialManager
            base = CredentialManager().retrieve("mantis_url") or ""
        except Exception:
            base = ""
        if not base:
            QMessageBox.warning(self, "未設定 Mantis URL", "請先至「系統設定」填寫 Mantis URL。")
            return
        # 萃取 base URL（去除頁面路徑，如 /view.php?id=...）
        try:
            from urllib.parse import urlparse
            parsed = urlparse(base.rstrip("/"))
            path = parsed.path
            if "." in path.rsplit("/", 1)[-1]:   # 末段含副檔名 → 去掉
                path = path[:path.rfind("/")]
            base_url = f"{parsed.scheme}://{parsed.netloc}{path}".rstrip("/")
        except Exception:
            base_url = base.rstrip("/")
        # 去除票號前導零以外的多餘格式（直接用原始字串）
        url = f"{base_url}/view.php?id={int(ticket_id)}"
        QDesktopServices.openUrl(QUrl(url))

    def _on_mark_pending_confirm(self) -> None:
        item = self._selected_item()
        if not item or not self._conn:
            return
        ReleaseManager(self._conn).mark_pending_confirm(item.id)
        self.refresh()
        label = item.mantis_ticket_id or item.case_id or f"#{item.id}"
        QMessageBox.information(self, "完成", f"已將 {label} 標記為「待確認」。")

    def _on_mark_released(self) -> None:
        item = self._selected_item()
        if not item or not self._conn:
            return
        ReleaseManager(self._conn).mark_released(item.id)
        self.refresh()
        label = item.mantis_ticket_id or item.case_id or f"#{item.id}"
        QMessageBox.information(self, "完成", f"已將 {label} 標記為已發布。")

    def _on_mark_pending(self) -> None:
        item = self._selected_item()
        if not item or not self._conn:
            return
        label = item.mantis_ticket_id or item.case_id or f"#{item.id}"
        reply = QMessageBox.question(
            self, "確認取消發布",
            f"將「{label}」狀態改回「待發」？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            ReleaseManager(self._conn).mark_pending(item.id)
            self.refresh()

    def _on_delete_item(self) -> None:
        item = self._selected_item()
        if not item or not self._conn:
            return
        label = item.mantis_ticket_id or item.case_id or f"#{item.id}"
        reply = QMessageBox.question(
            self, "確認刪除",
            f"確定刪除「{label}」？此操作無法復原。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            ReleaseManager(self._conn).delete_item(item.id)
            self.refresh()

    def _on_move_month(self) -> None:
        item = self._selected_item()
        if not item or not self._conn:
            return
        target_month = self._move_month_combo.currentData()
        if not target_month:
            return
        current_month = self._month_combo.currentData()
        if target_month == current_month:
            QMessageBox.information(self, "提示", "目標月份與目前月份相同，無需移動。")
            return
        label = item.mantis_ticket_id or item.case_id or f"#{item.id}"
        target_display = self._move_month_combo.currentText()
        reply = QMessageBox.question(
            self, "確認移動",
            f"將「{label}」從 {self._month_combo.currentText()} 移至 {target_display}？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            ReleaseManager(self._conn).update_month(item.id, target_month)
            self.refresh()
            QMessageBox.information(self, "完成", f"已將 {label} 移至 {target_display}。")
