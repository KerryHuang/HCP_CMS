"""KMS knowledge base view — search, CRUD, 待審核審核。"""

from __future__ import annotations

import re
import sqlite3
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from hcp_cms.data.models import QAKnowledge

from PySide6.QtCore import Qt, QUrl
from PySide6.QtGui import QPixmap
from PySide6.QtWebEngineWidgets import QWebEngineView
from PySide6.QtWidgets import (
    QDialog,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from hcp_cms.core.kms_engine import KMSEngine
from hcp_cms.ui.theme import ColorPalette, ThemeManager

_IMAGE_EXTS = frozenset({".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"})


class TextExpandDialog(QDialog):
    """欄位展開檢視對話框。"""

    def __init__(self, title: str, content: str, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(700, 500)
        layout = QVBoxLayout(self)
        text = QTextEdit()
        text.setReadOnly(True)
        text.setPlainText(content)
        layout.addWidget(text)
        close_btn = QPushButton("關閉")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)


class KMSImageViewDialog(QDialog):
    """完整回覆視窗：HTML + 圖片。"""

    def __init__(self, qa, db_dir: Path | None, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle(f"完整回覆 — {qa.qa_id}")
        self.resize(900, 700)
        self._qa = qa
        self._db_dir = db_dir
        self._setup_ui()

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)

        self._web = QWebEngineView()
        html = self._build_html()
        if self._db_dir:
            base_url = QUrl.fromLocalFile(str(self._db_dir) + "/")
            self._web.setHtml(html, base_url)
        else:
            self._web.setHtml(html)
        layout.addWidget(self._web, stretch=1)

        # 附件縮圖列
        img_dir = (self._db_dir / "kms_attachments" / self._qa.qa_id) if self._db_dir else None
        if img_dir and img_dir.exists():
            scroll = QScrollArea()
            scroll.setFixedHeight(120)
            scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
            scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            thumb_widget = QWidget()
            thumb_layout = QHBoxLayout(thumb_widget)
            for img_path in sorted(img_dir.iterdir()):
                if img_path.suffix.lower() in _IMAGE_EXTS:
                    lbl = QLabel()
                    pix = QPixmap(str(img_path)).scaled(
                        100,
                        100,
                        Qt.AspectRatioMode.KeepAspectRatio,
                        Qt.TransformationMode.SmoothTransformation,
                    )
                    lbl.setPixmap(pix)
                    lbl.setCursor(Qt.CursorShape.PointingHandCursor)
                    img_path_str = str(img_path)
                    lbl.mousePressEvent = lambda e, p=img_path_str: self._open_full_image(p)
                    thumb_layout.addWidget(lbl)
            thumb_layout.addStretch()
            scroll.setWidget(thumb_widget)
            layout.addWidget(scroll)

        close_btn = QPushButton("關閉")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)

    def _build_html(self) -> str:
        qa = self._qa
        # 嘗試重讀 .msg html_body
        html_body = None
        if qa.doc_name:
            msg_path = Path(qa.doc_name)
            if msg_path.exists():
                from hcp_cms.services.mail.msg_reader import MSGReader

                raw = MSGReader(directory=None).read_single_file(msg_path)
                if raw and raw.html_body:
                    html_body = self._replace_cid(raw.html_body)

        # 注入響應式 CSS，避免水平捲軸
        _RESPONSIVE_CSS = """<style>
body { max-width:100% !important; overflow-x:hidden !important;
       word-wrap:break-word !important; overflow-wrap:break-word !important; }
table { max-width:100% !important; table-layout:fixed !important; }
img { max-width:100% !important; height:auto !important; }
td, th { word-wrap:break-word !important; overflow-wrap:break-word !important; }
pre { white-space:pre-wrap !important; word-break:break-word !important; }
</style>"""

        if html_body:
            # 在 </head> 之前注入 CSS；若無 head 則插在最前面
            import re as _re2

            if _re2.search(r"</head>", html_body, _re2.IGNORECASE):
                html_body = _re2.sub(r"(</head>)", _RESPONSIVE_CSS + r"\1", html_body, count=1, flags=_re2.IGNORECASE)
            else:
                html_body = _RESPONSIVE_CSS + html_body
            return html_body

        # fallback：用圖片 + 文字組合簡易 HTML
        img_dir = (self._db_dir / "kms_attachments" / qa.qa_id) if self._db_dir else None
        img_html = ""
        if img_dir and img_dir.exists():
            for img_path in sorted(img_dir.iterdir()):
                if img_path.suffix.lower() in _IMAGE_EXTS:
                    img_html += f'<img src="{img_path.as_uri()}" style="max-width:100%;height:auto;margin:8px 0;"><br>'

        q = (qa.question or "").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        a = (qa.answer or "").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        s = (qa.solution or "").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
  body {{ font-family:sans-serif; padding:16px; max-width:100%;
          word-wrap:break-word; overflow-wrap:break-word; overflow-x:hidden; }}
  .content {{ line-height:1.6; }}
  img {{ max-width:100%; height:auto; }}
</style>
</head><body>
<h3>問題</h3><div class="content">{q}</div>
<h3>回覆</h3><div class="content">{a}</div>
{img_html}
{'<h3>解決方案</h3><div class="content">' + s + "</div>" if s else ""}
</body></html>"""

    def _replace_cid(self, html: str) -> str:
        """將 cid: 圖片參考替換為 file:// 路徑。"""
        if not self._db_dir:
            return html
        img_dir = self._db_dir / "kms_attachments" / self._qa.qa_id

        def replacer(m):
            cid = m.group(1).split("@")[0]
            for f in img_dir.iterdir() if img_dir.exists() else []:
                if f.stem.lower() == cid.lower() or f.name.lower() == cid.lower():
                    return f'src="{f.as_uri()}"'
            return m.group(0)

        return re.sub(r'src=["\']cid:([^"\'>\s]+)["\']', replacer, html, flags=re.IGNORECASE)

    def _open_full_image(self, img_path_str: str) -> None:
        dlg = QDialog(self)
        dlg.setWindowTitle("圖片檢視")
        dlg.resize(800, 600)
        lay = QVBoxLayout(dlg)
        lbl = QLabel()
        pix = QPixmap(img_path_str)
        lbl.setPixmap(
            pix.scaled(
                780,
                560,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
        )
        lay.addWidget(lbl)
        dlg.exec()


class QAReviewDialog(QDialog):
    """待審核 QA 編輯對話框。"""

    def __init__(self, qa, parent=None) -> None:
        super().__init__(parent)
        self.qa = qa
        self._result_action: str | None = None
        self.setWindowTitle(f"QA 審核編輯 — {qa.qa_id}")
        self.setMinimumWidth(600)
        self._setup_ui()

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        form = QFormLayout()

        self._question = QTextEdit(self.qa.question or "")
        self._question.setFixedHeight(80)
        self._answer = QTextEdit(self.qa.answer or "")
        self._answer.setFixedHeight(80)
        self._solution = QTextEdit(self.qa.solution or "")
        self._solution.setFixedHeight(60)
        self._keywords = QLineEdit(self.qa.keywords or "")
        self._product = QLineEdit(self.qa.system_product or "")

        form.addRow("問題：", self._question)
        form.addRow("回覆：", self._answer)
        form.addRow("解決方案：", self._solution)
        form.addRow("關鍵字：", self._keywords)
        form.addRow("產品：", self._product)
        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        draft_btn = QPushButton("💾 儲存草稿")
        draft_btn.clicked.connect(self._on_draft)
        approve_btn = QPushButton("✅ 確認完成")
        approve_btn.clicked.connect(self._on_approve)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(draft_btn)
        btn_layout.addWidget(approve_btn)
        layout.addLayout(btn_layout)

    def _collect_fields(self) -> dict:
        return {
            "question": self._question.toPlainText().strip(),
            "answer": self._answer.toPlainText().strip(),
            "solution": self._solution.toPlainText().strip() or None,
            "keywords": self._keywords.text().strip() or None,
            "system_product": self._product.text().strip() or None,
        }

    def _on_draft(self) -> None:
        self._result_action = "draft"
        self._result_fields = self._collect_fields()
        self.accept()

    def _on_approve(self) -> None:
        self._result_action = "approve"
        self._result_fields = self._collect_fields()
        self.accept()

    def result_action(self) -> str | None:
        return self._result_action

    def result_fields(self) -> dict:
        return getattr(self, "_result_fields", {})


class KMSView(QWidget):
    """KMS knowledge base page."""

    def __init__(
        self,
        conn: sqlite3.Connection | None = None,
        kms: KMSEngine | None = None,
        db_dir: Path | None = None,
        theme_mgr: ThemeManager | None = None,
    ) -> None:
        super().__init__()
        self._conn = conn
        self._kms = kms or (KMSEngine(conn) if conn else None)
        self._db_dir = db_dir
        self._theme_mgr = theme_mgr
        self._results: list = []
        self._pending: list = []
        self._setup_ui()
        if theme_mgr:
            self._apply_theme(theme_mgr.current_palette())
            theme_mgr.theme_changed.connect(self._apply_theme)

    def showEvent(self, event) -> None:
        """每次切換至 KMS 頁面時自動刷新待審核清單。"""
        super().showEvent(event)
        self._refresh_pending()

    def _make_field_widget(self, label: str, attr_name: str) -> tuple[QWidget, QTextEdit]:
        """建立標題列 + 展開按鈕 + QTextEdit 的組合 widget。"""
        wrapper = QWidget()
        vlay = QVBoxLayout(wrapper)
        vlay.setContentsMargins(0, 0, 0, 0)
        vlay.setSpacing(2)

        header = QWidget()
        hlay = QHBoxLayout(header)
        hlay.setContentsMargins(0, 0, 0, 0)
        lbl = QLabel(label)
        p = self._theme_mgr.current_palette() if self._theme_mgr else None
        lbl_color = p.text_tertiary if p else "#94a3b8"
        btn_color = p.text_muted if p else "#64748b"
        btn_hover = p.text_secondary if p else "#e2e8f0"
        lbl.setStyleSheet(f"color: {lbl_color}; font-size: 11px;")
        hlay.addWidget(lbl)
        hlay.addStretch()
        expand_btn = QPushButton("⛶")
        expand_btn.setFixedSize(22, 22)
        expand_btn.setToolTip("展開檢視")
        expand_btn.setStyleSheet(
            f"QPushButton {{ background: transparent; color: {btn_color}; font-size: 14px; border: none; }}"
            f" QPushButton:hover {{ color: {btn_hover}; }}"
        )
        hlay.addWidget(expand_btn)
        vlay.addWidget(header)

        edit = QTextEdit()
        edit.setReadOnly(True)
        vlay.addWidget(edit)

        field_label = label
        expand_btn.clicked.connect(
            lambda: TextExpandDialog(f"{field_label} — 展開檢視", edit.toPlainText(), self).exec()
        )
        return wrapper, edit

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)

        self._title = QLabel("📚 KMS 知識庫")
        layout.addWidget(self._title)

        self._tabs = QTabWidget()
        self._tabs.currentChanged.connect(self._on_tab_changed)

        # ── 全部 tab ──────────────────────────────────────────────────
        all_tab = QWidget()
        all_layout = QVBoxLayout(all_tab)

        # ── 搜尋列（獨立一排，輸入框放大）──────────────────────────
        search_row = QHBoxLayout()
        self._search_input = QLineEdit()
        self._search_input.setPlaceholderText("🔍  搜尋知識庫（支援同義詞擴展）...")
        self._search_input.setMinimumHeight(36)
        font = self._search_input.font()
        font.setPointSize(11)
        self._search_input.setFont(font)
        self._search_input.returnPressed.connect(self._on_search)
        search_row.addWidget(self._search_input, stretch=1)
        search_btn = QPushButton("🔍 搜尋")
        search_btn.setMinimumHeight(36)
        search_btn.setMinimumWidth(80)
        search_btn.clicked.connect(self._on_search)
        search_row.addWidget(search_btn)
        show_all_btn = QPushButton("📋 顯示全部")
        show_all_btn.setMinimumHeight(36)
        show_all_btn.clicked.connect(self._on_show_all)
        search_row.addWidget(show_all_btn)
        all_layout.addLayout(search_row)

        # ── 功能按鈕列（第二排）─────────────────────────────────────
        search_layout = QHBoxLayout()
        new_btn = QPushButton("➕ 新增 QA")
        new_btn.clicked.connect(self._on_new_qa)
        search_layout.addWidget(new_btn)
        import_msg_btn = QPushButton("📁 批次匯入 .msg")
        import_msg_btn.clicked.connect(self._on_import_msg_bulk)
        search_layout.addWidget(import_msg_btn)
        import_folder_btn = QPushButton("📂 匯入整個資料夾")
        import_folder_btn.clicked.connect(self._on_import_msg_folder)
        search_layout.addWidget(import_folder_btn)
        import_excel_btn = QPushButton("📊 匯入 Excel/CSV")
        import_excel_btn.clicked.connect(self._on_import_excel)
        search_layout.addWidget(import_excel_btn)
        export_sel_btn = QPushButton("💾 匯出選取")
        export_sel_btn.clicked.connect(self._on_export_selected)
        search_layout.addWidget(export_sel_btn)
        export_all_btn = QPushButton("📄 匯出全部")
        export_all_btn.clicked.connect(self._on_export_all)
        search_layout.addWidget(export_all_btn)
        search_layout.addStretch()
        all_layout.addLayout(search_layout)

        splitter = QSplitter(Qt.Orientation.Vertical)
        self._table = QTableWidget(0, 5)
        self._table.setHorizontalHeaderLabels(["QA 編號", "問題", "產品", "類型", "來源"])
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.itemSelectionChanged.connect(self._on_selection_changed)
        splitter.addWidget(self._table)

        detail = QWidget()
        detail_layout = QVBoxLayout(detail)
        detail_layout.setContentsMargins(4, 4, 4, 4)

        # 查看完整回覆按鈕（頂部）
        self._view_btn = QPushButton("🖼️ 查看完整回覆")
        self._view_btn.clicked.connect(self._on_view_full)
        detail_layout.addWidget(self._view_btn)

        # 三個欄位放入垂直 QSplitter
        field_splitter = QSplitter(Qt.Orientation.Vertical)

        q_widget, self._detail_question = self._make_field_widget("問題", "question")
        a_widget, self._detail_answer = self._make_field_widget("回覆", "answer")
        s_widget, self._detail_solution = self._make_field_widget("解決方案", "solution")

        field_splitter.addWidget(q_widget)
        field_splitter.addWidget(a_widget)
        field_splitter.addWidget(s_widget)
        field_splitter.setSizes([200, 400, 200])

        detail_layout.addWidget(field_splitter)

        # 單筆匯出按鈕
        self._export_single_btn = QPushButton("💾 另存 .docx")
        self._export_single_btn.clicked.connect(self._on_export_single)
        detail_layout.addWidget(self._export_single_btn)

        splitter.addWidget(detail)
        all_layout.addWidget(splitter)
        self._tabs.addTab(all_tab, "全部")

        # ── 待審核 tab ────────────────────────────────────────────────
        pending_tab = QWidget()
        pending_layout = QVBoxLayout(pending_tab)

        self._pending_table = QTableWidget(0, 3)
        self._pending_table.setHorizontalHeaderLabels(["QA 編號", "問題預覽", "來源案件"])
        self._pending_table.horizontalHeader().setStretchLastSection(True)
        self._pending_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._pending_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._pending_table.doubleClicked.connect(self._on_review)
        pending_layout.addWidget(self._pending_table)

        pending_btn_layout = QHBoxLayout()
        review_btn = QPushButton("✏️ 編輯審核")
        review_btn.clicked.connect(self._on_review)
        delete_btn = QPushButton("🗑️ 刪除")
        delete_btn.clicked.connect(self._on_delete_pending)
        pending_btn_layout.addWidget(review_btn)
        pending_btn_layout.addWidget(delete_btn)
        pending_btn_layout.addStretch()
        pending_layout.addLayout(pending_btn_layout)
        self._tabs.addTab(pending_tab, "待審核")

        # ── 智慧查詢 tab ──────────────────────────────────────────────
        smart_tab = QWidget()
        smart_layout = QVBoxLayout(smart_tab)
        smart_layout.setSpacing(8)

        self._smart_desc = QLabel(
            "📋 將客戶來信的問題描述貼入下方，系統會自動分析語意、比對歷史 Q&A，整理出建議回覆供參考使用。"
        )
        self._smart_desc.setWordWrap(True)
        smart_layout.addWidget(self._smart_desc)

        self._smart_input = QTextEdit()
        self._smart_input.setPlaceholderText("請在此貼上客戶問題原文（可貼整封信件，系統自動萃取關鍵詞比對）...")
        self._smart_input.setMinimumHeight(120)
        self._smart_input.setMaximumHeight(180)
        smart_layout.addWidget(self._smart_input)

        smart_btn_row = QHBoxLayout()
        smart_search_btn = QPushButton("🧠 分析比對")
        smart_search_btn.setMinimumHeight(36)
        smart_search_btn.setMinimumWidth(120)
        smart_search_btn.clicked.connect(self._on_smart_search)
        smart_btn_row.addWidget(smart_search_btn)
        smart_clear_btn = QPushButton("🗑 清除")
        smart_clear_btn.setMinimumHeight(36)
        smart_clear_btn.clicked.connect(
            lambda: (
                self._smart_input.clear(),
                self._smart_result_table.setRowCount(0),
                self._smart_answer_box.clear(),
            )
        )
        smart_btn_row.addWidget(smart_clear_btn)
        smart_btn_row.addStretch()
        smart_layout.addLayout(smart_btn_row)

        self._smart_result_label = QLabel("比對結果（點選查看建議回覆）：")
        smart_layout.addWidget(self._smart_result_label)

        smart_splitter = QSplitter(Qt.Orientation.Vertical)

        self._smart_result_table = QTableWidget(0, 4)
        self._smart_result_table.setHorizontalHeaderLabels(["相符度", "QA 編號", "問題摘要", "產品"])
        self._smart_result_table.horizontalHeader().setStretchLastSection(True)
        self._smart_result_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._smart_result_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._smart_result_table.itemSelectionChanged.connect(self._on_smart_select)
        self._smart_result_table.setColumnWidth(0, 70)
        self._smart_result_table.setColumnWidth(1, 120)
        self._smart_result_table.setColumnWidth(2, 400)
        smart_splitter.addWidget(self._smart_result_table)

        answer_widget = QWidget()
        answer_layout = QVBoxLayout(answer_widget)
        answer_layout.setContentsMargins(0, 0, 0, 0)
        self._answer_label = QLabel("建議回覆：")
        answer_layout.addWidget(self._answer_label)
        self._smart_answer_box = QTextEdit()
        self._smart_answer_box.setReadOnly(True)
        answer_layout.addWidget(self._smart_answer_box)
        smart_splitter.addWidget(answer_widget)
        smart_splitter.setSizes([200, 300])

        smart_layout.addWidget(smart_splitter, stretch=1)
        self._tabs.addTab(smart_tab, "🧠 智慧查詢")

        self._smart_results: list[QAKnowledge] = []

        layout.addWidget(self._tabs)

    def _on_tab_changed(self, index: int) -> None:
        if index == 1:
            self._refresh_pending()

    def _refresh_pending(self) -> None:
        if not self._kms:
            return
        self._pending = self._kms.list_pending()
        count = len(self._pending)
        self._tabs.setTabText(1, f"待審核{'  🔴' + str(count) if count else ''}")
        self._pending_table.setRowCount(count)
        for i, qa in enumerate(self._pending):
            self._pending_table.setItem(i, 0, QTableWidgetItem(qa.qa_id))
            self._pending_table.setItem(i, 1, QTableWidgetItem((qa.question or "")[:50]))
            self._pending_table.setItem(i, 2, QTableWidgetItem(qa.source_case_id or ""))

    def _on_show_all(self) -> None:
        """載入所有已完成 QA 顯示於全部 tab。"""
        if not self._kms:
            return
        self._results = self._kms.list_approved()
        self._table.setRowCount(len(self._results))
        for i, qa in enumerate(self._results):
            self._table.setItem(i, 0, QTableWidgetItem(qa.qa_id))
            self._table.setItem(i, 1, QTableWidgetItem(qa.question or ""))
            self._table.setItem(i, 2, QTableWidgetItem(qa.system_product or ""))
            self._table.setItem(i, 3, QTableWidgetItem(qa.issue_type or ""))
            self._table.setItem(i, 4, QTableWidgetItem(qa.source or ""))

    def _on_search(self) -> None:
        query = self._search_input.text().strip()
        if not query or not self._kms:
            return
        self._results = self._kms.search(query)
        self._table.setRowCount(len(self._results))
        for i, qa in enumerate(self._results):
            self._table.setItem(i, 0, QTableWidgetItem(qa.qa_id))
            self._table.setItem(i, 1, QTableWidgetItem(qa.question or ""))
            self._table.setItem(i, 2, QTableWidgetItem(qa.system_product or ""))
            self._table.setItem(i, 3, QTableWidgetItem(qa.issue_type or ""))
            self._table.setItem(i, 4, QTableWidgetItem(qa.source or ""))

    def _on_selection_changed(self) -> None:
        rows = self._table.selectionModel().selectedRows()
        if not rows or not self._results:
            return
        row = rows[0].row()
        if row < 0 or row >= len(self._results):
            return
        qa = self._results[row]
        self._detail_question.setPlainText(qa.question or "")
        self._detail_answer.setPlainText(qa.answer or "")
        self._detail_solution.setPlainText(qa.solution or "")
        # 高亮查看按鈕
        if qa.has_image == "是":
            accent = self._theme_mgr.current_palette().accent if self._theme_mgr else "#60a5fa"
            self._view_btn.setStyleSheet(f"color: {accent};")
        else:
            self._view_btn.setStyleSheet("")

    def _on_new_qa(self) -> None:
        """手動新增 QA 對話框。"""
        if not self._kms:
            return
        from hcp_cms.data.models import QAKnowledge

        empty = QAKnowledge(qa_id="（新增）", question="", answer="")
        dlg = QAReviewDialog(empty, parent=self)
        dlg.setWindowTitle("新增 QA")
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        fields = dlg.result_fields()
        if not fields.get("question") or not fields.get("answer"):
            from PySide6.QtWidgets import QMessageBox

            QMessageBox.warning(self, "欄位不完整", "「問題」與「回覆」為必填欄位。")
            return
        action = dlg.result_action()
        status = "已完成" if action == "approve" else "待審核"
        self._kms.create_qa(status=status, **fields)
        self._refresh_pending()
        self._on_show_all()

    def _on_import_msg_bulk(self) -> None:
        """批次選取多個 .msg 直接建立 KMS QA（待審核）。"""
        if not self._kms:
            return
        from PySide6.QtWidgets import QMessageBox, QProgressDialog

        from hcp_cms.services.mail.msg_reader import MSGReader

        files, _ = QFileDialog.getOpenFileNames(self, "選擇 .msg 檔案（可多選）", "", "Outlook Messages (*.msg)")
        if not files:
            return

        paths = [Path(f) for f in files]
        reader = MSGReader()

        progress = QProgressDialog("批次匯入中...", "取消", 0, len(paths), self)
        progress.setWindowTitle("批次匯入 .msg")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.show()

        ok, fail = 0, 0
        for i, path in enumerate(paths):
            if progress.wasCanceled():
                break
            progress.setLabelText(f"處理 {path.name}（{i + 1}/{len(paths)}）")
            progress.setValue(i)
            try:
                email = reader.read_single_file(path)
                if email:
                    self._kms.extract_qa_from_email(email, db_dir=self._db_dir)
                    ok += 1
                else:
                    fail += 1
            except Exception:
                fail += 1
        progress.setValue(len(paths))

        self._refresh_pending()
        self._on_show_all()
        QMessageBox.information(
            self, "批次匯入完成", f"✅ 成功：{ok} 筆\n❌ 失敗／略過：{fail} 筆\n\n已匯入的 QA 位於「待審核」頁籤。"
        )

    def _on_import_msg_folder(self) -> None:
        """選擇資料夾，自動遞迴讀取所有 .msg / .xlsx / .xls / .csv 並建立 KMS QA。"""
        if not self._kms:
            return
        import csv

        import openpyxl
        from PySide6.QtWidgets import QMessageBox, QProgressDialog

        from hcp_cms.services.mail.msg_reader import MSGReader

        folder = QFileDialog.getExistingDirectory(self, "選擇資料夾（含子資料夾，支援 .msg / .xlsx / .xls / .csv）", "")
        if not folder:
            return

        root = Path(folder)
        msg_paths = sorted(root.rglob("*.msg"))
        excel_paths = sorted(root.rglob("*.xlsx")) + sorted(root.rglob("*.xls")) + sorted(root.rglob("*.csv"))
        all_paths = msg_paths + excel_paths

        if not all_paths:
            QMessageBox.information(
                self, "無檔案", f"在資料夾中找不到任何支援的檔案：\n{folder}\n\n支援格式：.msg、.xlsx、.xls、.csv"
            )
            return

        reader = MSGReader()
        progress = QProgressDialog("批次匯入中...", "取消", 0, len(all_paths), self)
        progress.setWindowTitle(f"匯入資料夾（.msg×{len(msg_paths)} + Excel/CSV×{len(excel_paths)}）")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.show()

        COL_Q = ["問題", "question", "Q", "題目", "issue"]
        COL_A = ["回覆", "答案", "answer", "A", "回答", "reply"]
        COL_S = ["解決方案", "solution", "解法"]
        COL_K = ["關鍵字", "keywords", "keyword", "關鍵詞"]
        COL_P = ["產品", "product", "系統", "system_product"]

        def _find(row: dict, keys: list[str]) -> str | None:
            for k in keys:
                if k in row and row[k]:
                    return row[k]
            return None

        def _import_excel_rows(rows: list[dict]) -> tuple[int, int]:
            ok2, fail2 = 0, 0
            for row in rows:
                q = _find(row, COL_Q)
                a = _find(row, COL_A)
                if not q and not a:
                    continue  # 完全空白列，靜默略過
                if not q or not a:
                    fail2 += 1
                    continue
                try:
                    self._kms.create_qa(
                        question=q,
                        answer=a,
                        solution=_find(row, COL_S),
                        keywords=_find(row, COL_K),
                        system_product=_find(row, COL_P),
                        status="已完成",
                    )
                    ok2 += 1
                except Exception:
                    fail2 += 1
            return ok2, fail2

        ok, fail = 0, 0
        for i, path in enumerate(all_paths):
            if progress.wasCanceled():
                break
            progress.setLabelText(f"處理 {path.name}（{i + 1}/{len(all_paths)}）")
            progress.setValue(i)
            try:
                if path.suffix.lower() == ".msg":
                    email = reader.read_single_file(path)
                    if email:
                        self._kms.extract_qa_from_email(email, db_dir=self._db_dir)
                        ok += 1
                    else:
                        fail += 1
                elif path.suffix.lower() in (".xlsx", ".xls"):
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    header_row = next(ws.iter_rows(max_row=1))
                    headers = [str(c.value or "").strip() for c in header_row]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, [str(v or "").strip() for v in row])))
                    wb.close()
                    o, f = _import_excel_rows(rows)
                    ok += o
                    fail += f
                else:  # .csv
                    with open(path, newline="", encoding="utf-8-sig") as f_csv:
                        rows = [{k.strip(): (v or "").strip() for k, v in r.items()} for r in csv.DictReader(f_csv)]
                    o, f = _import_excel_rows(rows)
                    ok += o
                    fail += f
            except Exception:
                fail += 1
        progress.setValue(len(all_paths))

        self._refresh_pending()
        self._on_show_all()
        QMessageBox.information(
            self,
            "資料夾匯入完成",
            f"資料夾：{folder}\n\n"
            f"📧 .msg 檔案：{len(msg_paths)} 個\n"
            f"📊 Excel/CSV 檔案：{len(excel_paths)} 個\n\n"
            f"✅ 成功：{ok} 筆\n❌ 失敗／略過：{fail} 筆\n\n"
            ".msg → 「待審核」頁籤；Excel/CSV → 直接「已完成」",
        )

    def _on_import_excel(self) -> None:
        """從 Excel / CSV 批次匯入 QA（直接建立為「已完成」）。"""
        if not self._kms:
            return
        import csv

        import openpyxl
        from PySide6.QtWidgets import QMessageBox

        path, _ = QFileDialog.getOpenFileName(self, "選擇 Excel 或 CSV 檔案", "", "試算表 (*.xlsx *.xls *.csv)")
        if not path:
            return

        p = Path(path)
        rows: list[dict] = []
        try:
            if p.suffix.lower() in (".xlsx", ".xls"):
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                ws = wb.active
                header_row = next(ws.iter_rows(max_row=1))
                headers = [str(c.value or "").strip() for c in header_row]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v or "").strip() for v in row])))
                wb.close()
            else:
                with open(p, newline="", encoding="utf-8-sig") as f:
                    for row in csv.DictReader(f):
                        rows.append({k.strip(): (v or "").strip() for k, v in row.items()})
        except Exception as e:
            QMessageBox.critical(self, "讀取失敗", f"無法解析檔案：\n{e}")
            return

        # 彈性欄位對應
        COL_Q = ["問題", "question", "Q", "題目", "issue"]
        COL_A = ["回覆", "答案", "answer", "A", "回答", "reply"]
        COL_S = ["解決方案", "solution", "解法"]
        COL_K = ["關鍵字", "keywords", "keyword", "關鍵詞"]
        COL_P = ["產品", "product", "系統", "system_product"]

        def _find(row: dict, candidates: list[str]) -> str | None:
            for c in candidates:
                if c in row and row[c]:
                    return row[c]
            return None

        ok, fail = 0, 0
        for row in rows:
            q = _find(row, COL_Q)
            a = _find(row, COL_A)
            if not q and not a:
                continue  # 完全空白列，靜默略過
            if not q or not a:
                fail += 1
                continue
            try:
                self._kms.create_qa(
                    question=q,
                    answer=a,
                    solution=_find(row, COL_S),
                    keywords=_find(row, COL_K),
                    system_product=_find(row, COL_P),
                    status="已完成",
                )
                ok += 1
            except Exception:
                fail += 1

        self._refresh_pending()
        self._on_show_all()
        QMessageBox.information(
            self,
            "匯入完成",
            f"✅ 成功：{ok} 筆\n❌ 失敗／略過：{fail} 筆\n\n"
            "📋 Excel 欄位名稱（擇一）：\n"
            "  問題（必填）、回覆（必填）、解決方案、關鍵字、產品",
        )

    def _on_smart_search(self) -> None:
        """智慧查詢：貼入客戶問題原文，比對 KMS 知識庫，回傳最相近的 QA。"""
        if not self._kms:
            return
        query = self._smart_input.toPlainText().strip()
        if not query:
            return

        from PySide6.QtWidgets import QTableWidgetItem

        results = self._kms.search(query)
        self._smart_results = results[:10]
        self._smart_result_table.setRowCount(0)
        self._smart_answer_box.clear()

        if not results:
            self._smart_answer_box.setPlainText("⚠️ 未找到相符的歷史 Q&A，建議手動回覆。")
            return

        for rank, qa in enumerate(self._smart_results, start=1):
            row = self._smart_result_table.rowCount()
            self._smart_result_table.insertRow(row)
            pct = max(0, 100 - (rank - 1) * 8)
            self._smart_result_table.setItem(row, 0, QTableWidgetItem(f"{pct}%"))
            self._smart_result_table.setItem(row, 1, QTableWidgetItem(qa.qa_id))
            self._smart_result_table.setItem(row, 2, QTableWidgetItem(qa.question[:80] if qa.question else ""))
            self._smart_result_table.setItem(row, 3, QTableWidgetItem(qa.system_product or ""))

        # 預設選第一筆
        self._smart_result_table.selectRow(0)

    def _on_smart_select(self) -> None:
        """點選智慧查詢結果，顯示建議回覆。"""
        rows = self._smart_result_table.selectionModel().selectedRows()
        if not rows or not self._smart_results:
            return
        idx = rows[0].row()
        if idx >= len(self._smart_results):
            return
        qa = self._smart_results[idx]
        parts = []
        if qa.answer:
            parts.append(f"【建議回覆】\n{qa.answer}")
        if qa.solution:
            parts.append(f"\n【解決方案】\n{qa.solution}")
        parts.append(f"\n— 來源：{qa.qa_id}")
        self._smart_answer_box.setPlainText("\n".join(parts))

    def _on_review(self) -> None:
        if not self._kms:
            return
        rows = self._pending_table.selectionModel().selectedRows()
        if not rows:
            return
        row = rows[0].row()
        if row >= len(self._pending):
            return
        qa = self._pending[row]
        dlg = QAReviewDialog(qa, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        fields = dlg.result_fields()
        if dlg.result_action() == "draft":
            self._kms.update_qa(qa.qa_id, **fields)
        elif dlg.result_action() == "approve":
            self._kms.approve_qa(qa.qa_id, **fields)
        self._refresh_pending()

    def _on_delete_pending(self) -> None:
        if not self._kms:
            return
        rows = self._pending_table.selectionModel().selectedRows()
        if not rows:
            return
        row = rows[0].row()
        if row >= len(self._pending):
            return
        qa = self._pending[row]
        self._kms.delete_qa(qa.qa_id)
        self._refresh_pending()

    def _on_view_full(self) -> None:
        rows = self._table.selectionModel().selectedRows()
        if not rows or not self._results:
            return
        qa = self._results[rows[0].row()]
        dlg = KMSImageViewDialog(qa, self._db_dir, parent=self)
        dlg.exec()

    def _on_export_single(self) -> None:
        rows = self._table.selectionModel().selectedRows()
        if not rows or not self._results:
            return
        qa = self._results[rows[0].row()]
        self._do_export([qa])

    def _on_export_selected(self) -> None:
        rows = self._table.selectionModel().selectedRows()
        if not rows:
            return
        qa_list = [self._results[r.row()] for r in rows if r.row() < len(self._results)]
        self._do_export(qa_list)

    def _on_export_all(self) -> None:
        self._do_export(self._results if self._results else [])

    def _do_export(self, qa_list: list) -> None:
        if not self._kms:
            return
        today = date.today().strftime("%Y%m%d")
        default_name = f"KMS匯出_{today}.docx"
        path, _ = QFileDialog.getSaveFileName(self, "另存 Word 文件", default_name, "Word 文件 (*.docx)")
        if not path:
            return
        try:
            self._kms.export_to_docx(Path(path), db_dir=self._db_dir or Path("."), qa_list=qa_list)
            QMessageBox.information(self, "匯出完成", f"已儲存至：\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "匯出失敗", str(e))

    def _apply_theme(self, p: ColorPalette) -> None:
        """套用主題色彩。"""
        self._title.setStyleSheet(f"font-size: 18px; font-weight: bold; color: {p.text_primary};")
        self._smart_desc.setStyleSheet(f"color: {p.text_tertiary}; padding: 4px 0;")
        self._smart_result_label.setStyleSheet(f"color: {p.text_tertiary}; font-size: 11px;")
        self._answer_label.setStyleSheet(f"color: {p.accent}; font-weight: bold;")
        self._smart_answer_box.setStyleSheet(
            f"background: {p.bg_secondary}; color: {p.text_primary}; "
            f"border: 1px solid {p.border_primary}; border-radius: 4px;"
        )
