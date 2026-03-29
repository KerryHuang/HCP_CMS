"""Excel 匯出工具 — 寄件備份。"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from hcp_cms.core.sent_mail_manager import EnrichedSentMail


class ExcelExporter:
    """將 EnrichedSentMail 列表匯出為 xlsx（兩個工作表）。"""

    def export_sent_mail(self, mails: list[EnrichedSentMail], path: str) -> None:
        """匯出寄件備份至 xlsx。

        Args:
            mails: 已抓取並豐富化的寄件清單。
            path: 輸出檔案路徑（含副檔名 .xlsx）。
        """
        wb = Workbook()
        self._write_summary(wb, mails)
        self._write_list(wb, mails)
        # 刪除預設空白工作表
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(path)

    def _write_summary(self, wb: Workbook, mails: list[EnrichedSentMail]) -> None:
        ws = wb.create_sheet("公司彙總")
        bold = Font(bold=True)

        # 標題列
        for col, title in enumerate(["公司名稱", "次數"], start=1):
            cell = ws.cell(1, col, title)
            cell.font = bold

        # 去重並依次數降冪
        seen: dict[str, tuple[str, int]] = {}
        for m in mails:
            if m.company_id and m.company_id not in seen:
                seen[m.company_id] = (m.company_name or m.company_id, m.company_reply_count)
        ranked = sorted(seen.values(), key=lambda x: x[1], reverse=True)

        for row, (name, count) in enumerate(ranked, start=2):
            ws.cell(row, 1, name)
            ws.cell(row, 2, count)

    def _write_list(self, wb: Workbook, mails: list[EnrichedSentMail]) -> None:
        ws = wb.create_sheet("寄件清單")
        bold = Font(bold=True)

        headers = ["日期", "收件人", "主旨", "公司", "案件", "第幾封"]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(1, col, title)
            cell.font = bold

        company_counters: dict[str, int] = {}
        for row, m in enumerate(mails, start=2):
            ws.cell(row, 1, m.date if m.date else "")
            ws.cell(row, 2, ", ".join(m.recipients))
            ws.cell(row, 3, m.subject)
            ws.cell(row, 4, m.company_name or "—")
            ws.cell(row, 5, m.linked_case_id or "—")
            if m.company_id:
                company_counters[m.company_id] = company_counters.get(m.company_id, 0) + 1
                ws.cell(row, 6, str(company_counters[m.company_id]))
            else:
                ws.cell(row, 6, "—")
