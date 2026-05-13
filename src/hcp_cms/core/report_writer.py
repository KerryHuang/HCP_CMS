"""Excel report writer — writes structured data to styled .xlsx files."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.hyperlink import Hyperlink


@dataclass
class HyperlinkCell:
    """Excel 內部超連結格子：顯示 text，點擊跳轉至同檔案的 target_sheet 工作表。"""

    text: str
    target_sheet: str

    def __str__(self) -> str:
        return self.text

FONT_HEADER = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
FILL_ALT_ROW = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Mantis 追蹤工作表分類色彩（背景色, 字體色）
_MANTIS_COLORS: dict[str, tuple[str, str]] = {
    "high":   ("FECACA", "7F1D1D"),
    "salary": ("FEF3C7", "78350F"),
    "normal": ("FFFFFF", "111827"),
    "closed": ("F3F4F6", "6B7280"),
}

# 美化樣式（B 方案 + 丙：標題、分隔列、handler 分色、超時分級）

# 大標題列（單元格的 sheet title 列）
FONT_TITLE = Font(name="微軟正黑體", size=14, bold=True, color="1E3A5F")
FILL_TITLE = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

# Numbered section 標題（1. 處理中案件...）
FONT_SECTION = Font(name="微軟正黑體", size=12, bold=True, color="111827")
FILL_SECTION = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

# Handler 分色（各客服一個色）
_HANDLER_FILLS: dict[str, str] = {
    "jill": "DBEAFE",       # 淡藍
    "yoga": "D1FAE5",       # 淡綠
    "rebecca": "FCE7F3",    # 淡粉紅
    "（未指派）": "FEF3C7",  # 淡黃
}
_HANDLER_FALLBACK_PALETTE = ["E9D5FF", "FED7AA", "BFDBFE", "FEE2E2", "CCFBF1"]
FONT_HANDLER_HEADER = Font(name="微軟正黑體", size=12, bold=True, color="111827")

# 超時 5 級背景色（與 case_view._overdue_color 對齊，但用較淡色適合 Excel 白底）
OVERDUE_FILLS: list[tuple[int, str]] = [
    (30, "FECACA"),   # 紅
    (10, "FED7AA"),   # 橘
    (7, "FEF08A"),    # 黃
    (5, "FEF3C7"),    # 淡黃
    (3, "FFFBEB"),    # 米黃
]


def _is_title_row(rows: list[list], row_idx: int) -> bool:
    """判斷是否為大標題列（第 1 列且只佔 1 格、含表情符號或破折號標識）。"""
    if row_idx != 0:
        return False
    if not rows or not rows[0] or len(rows[0]) != 1:
        return False
    s = str(rows[0][0])
    return any(c in s for c in "📊📋🏢⏰👥📌📈🚨❓") or "—" in s


def _is_section_divider(row: list) -> bool:
    """━━ jill — ... ━━ 這種 handler section 分隔列。"""
    return len(row) == 1 and row[0] and "━━" in str(row[0])


def _is_numbered_section(row: list) -> bool:
    """1. 處理中... / 2. 各處理人... 這種編號 section 標題。"""
    if len(row) != 1 or not row[0]:
        return False
    s = str(row[0]).strip()
    return len(s) > 2 and s[0].isdigit() and s[1] == "."


def _extract_handler_from_divider(s: str) -> str:
    """從「━━ jill — ... ━━」抽取 handler 名稱（小寫）。"""
    import re
    m = re.search(r"━━\s*(\S+?)\s*—", s)
    return m.group(1).lower() if m else ""


def _handler_fill_color(handler_name_lower: str, counter: dict[str, int]) -> str:
    """取得 handler 的背景色（已知 handler 用固定色，其他用 fallback palette）。"""
    if handler_name_lower in _HANDLER_FILLS:
        return _HANDLER_FILLS[handler_name_lower]
    # 未指派
    if handler_name_lower in ("（未指派）", "未指派"):
        return _HANDLER_FILLS["（未指派）"]
    # Fallback：依 counter 輪替
    idx = counter.setdefault(handler_name_lower, len(counter))
    return _HANDLER_FALLBACK_PALETTE[idx % len(_HANDLER_FALLBACK_PALETTE)]


def _overdue_fill_color(days_stuck: int) -> str | None:
    """依卡幾天回傳 Excel 背景色（< 3 天回傳 None）。"""
    for threshold, color in OVERDUE_FILLS:
        if days_stuck >= threshold:
            return color
    return None


class ReportWriter:
    """Writes structured report data to Excel files with styling."""

    @staticmethod
    def write_excel(data: dict[str, list[list]], path: Path) -> None:
        """Write structured data to an Excel file with enhanced styling.

        美化規則：
        - 標題列（第 1 列且單格、含表情符號 / 破折號）：大字 + 淡藍底 + 合併
        - Section 分隔列（含「━━」）：合併 + handler 分色背景
        - Numbered section（如「1. ...」）：合併 + 灰色背景 + 粗體
        - 資料列若有「卡幾天」欄且數值 >= 3：依 5 級漸層上色
        - 凍結首列

        Args:
            data: dict mapping sheet_name to rows. First row of each sheet = header.
            path: Output file path.
        """
        wb = openpyxl.Workbook()
        first = True

        for sheet_name, rows in data.items():
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(sheet_name)

            if not rows:
                continue

            # 全 sheet 最大欄數，供合併用
            max_cols = max((len(r) for r in rows), default=1) or 1
            handler_counter: dict[str, int] = {}
            # 紀錄目前 section 的 handler 色（資料列使用）
            current_section_fill: str | None = None
            # 第一個「正規 header」（有 col_count >= 2 的）row index，用於凍結
            first_data_header_idx: int | None = None
            # 「卡幾天」欄索引（每次 header 重設）
            days_col_idx: int | None = None

            for row_idx, row in enumerate(rows, 1):
                if not row:
                    continue

                # ── 大標題列（單格、含表情符號）─────────────────────
                if _is_title_row(rows, row_idx - 1):
                    cell = ws.cell(row=row_idx, column=1, value=str(row[0]))
                    cell.font = FONT_TITLE
                    cell.fill = FILL_TITLE
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if max_cols > 1:
                        ws.merge_cells(
                            start_row=row_idx, start_column=1,
                            end_row=row_idx, end_column=max_cols,
                        )
                    ws.row_dimensions[row_idx].height = 24
                    continue

                # ── Section 分隔列（━━ jill — ... ━━）─────────────
                if _is_section_divider(row):
                    s = str(row[0])
                    handler_lc = _extract_handler_from_divider(s)
                    fill_color = _handler_fill_color(handler_lc, handler_counter)
                    current_section_fill = fill_color
                    cell = ws.cell(row=row_idx, column=1, value=s)
                    cell.font = FONT_HANDLER_HEADER
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    # 多行內容需開啟 wrap_text；垂直置中
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    if max_cols > 1:
                        ws.merge_cells(
                            start_row=row_idx, start_column=1,
                            end_row=row_idx, end_column=max_cols,
                        )
                    # 依行數調整列高（每行約 18px）
                    line_count = s.count("\n") + 1
                    ws.row_dimensions[row_idx].height = max(22, line_count * 20)
                    continue

                # ── 編號 section 標題（1./2./3.）─────────────────────
                if _is_numbered_section(row):
                    current_section_fill = None  # 重設 handler 色
                    cell = ws.cell(row=row_idx, column=1, value=str(row[0]))
                    cell.font = FONT_SECTION
                    cell.fill = FILL_SECTION
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if max_cols > 1:
                        ws.merge_cells(
                            start_row=row_idx, start_column=1,
                            end_row=row_idx, end_column=max_cols,
                        )
                    ws.row_dimensions[row_idx].height = 20
                    continue

                # ── Header 列（第一個非特殊列、或剛剛遇到 section 後的第一個一般列）─
                # 在 sheet 內，每個 section 之後第一個「多格」列當 header
                is_first_row_of_sheet = first_data_header_idx is None and row_idx == 1
                # 簡化：sheet 第一列若非特殊則一定是 header
                # 後續若想偵測新 section 的 header，可由 numbered_section 重設
                if is_first_row_of_sheet or (first_data_header_idx is None):
                    first_data_header_idx = row_idx
                    days_col_idx = None
                    for col, value in enumerate(row, 1):
                        if str(value) == "卡幾天":
                            days_col_idx = col
                        if isinstance(value, HyperlinkCell):
                            cell = ws.cell(row=row_idx, column=col, value=value.text)
                            escaped = value.target_sheet.replace("'", "''")
                            cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{escaped}'!A1")
                            cell.font = Font(name="微軟正黑體", size=11, bold=True, color="0563C1", underline="single")
                        else:
                            cell = ws.cell(row=row_idx, column=col, value=value)
                            cell.font = FONT_HEADER
                            cell.fill = FILL_HEADER
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = BORDER_THIN
                    continue

                # Section 後的子表頭（如「案件編號 公司 主旨 卡幾天 回覆次數」）
                # 偵測：列內 >= 2 格、所有格都是非空字串
                is_subheader = (
                    len(row) >= 2
                    and all(isinstance(v, str) and v for v in row)
                )
                if is_subheader:
                    days_col_idx = None
                    for col, value in enumerate(row, 1):
                        if str(value) == "卡幾天":
                            days_col_idx = col
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.font = FONT_HEADER
                        cell.fill = FILL_HEADER
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = BORDER_THIN
                    continue

                # ── 一般資料列 ─────────────────────────────────────
                # 判斷是否要套用超時色彩
                overdue_fill: str | None = None
                if days_col_idx is not None and days_col_idx <= len(row):
                    raw_days = row[days_col_idx - 1]
                    if isinstance(raw_days, int):
                        overdue_fill = _overdue_fill_color(raw_days)

                for col, value in enumerate(row, 1):
                    if isinstance(value, HyperlinkCell):
                        cell = ws.cell(row=row_idx, column=col, value=value.text)
                        escaped = value.target_sheet.replace("'", "''")
                        cell.hyperlink = Hyperlink(
                            ref=cell.coordinate, location=f"'{escaped}'!A1",
                        )
                        cell.font = Font(name="微軟正黑體", size=10, color="0563C1", underline="single")
                    else:
                        cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = BORDER_THIN
                    # 優先順序：超時 > handler section > 交替
                    if overdue_fill and not isinstance(value, HyperlinkCell):
                        cell.fill = PatternFill(
                            start_color=overdue_fill, end_color=overdue_fill, fill_type="solid")
                    elif current_section_fill and not isinstance(value, HyperlinkCell):
                        # 較淡的 section 色（資料列）
                        cell.fill = PatternFill(
                            start_color=current_section_fill,
                            end_color=current_section_fill, fill_type="solid")
                    elif row_idx % 2 == 0 and not isinstance(value, HyperlinkCell):
                        cell.fill = FILL_ALT_ROW

            # 凍結首個 header 列下方
            if first_data_header_idx:
                ws.freeze_panes = f"A{first_data_header_idx + 1}"

        # 自動調整欄寬（依內容最大長度，上限 50 字元）
        for ws in wb.worksheets:
            for col_cells in ws.columns:
                max_length = 0
                for cell in col_cells:
                    # 合併儲存格的「跟隨格」會是 MergedCell 沒 value，略過
                    if hasattr(cell, "value") and cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted = min(max_length + 2, 50)
                if adjusted > 0:
                    # 跳過合併儲存格的副欄
                    try:
                        col_letter = col_cells[0].column_letter
                        ws.column_dimensions[col_letter].width = adjusted
                    except AttributeError:
                        continue

        wb.save(str(path))

    @staticmethod
    def append_mantis_sheet(
        path: Path,
        sheet_name: str,
        rows: list[dict],
    ) -> None:
        """在已存在的 Excel 檔案中新增 Mantis 追蹤工作表（帶分色）。

        Args:
            path:       已存在的 .xlsx 檔案路徑。
            sheet_name: 工作表名稱，如「📌 Mantis 追蹤」。
            rows:       build_mantis_sheet() 回傳的 list[dict]，
                        每列需含 category 欄位。
        """
        headers = ["#", "票號", "摘要", "狀態", "優先", "未處理天數", "最後更新", "負責人"]

        wb = openpyxl.load_workbook(str(path))
        ws = wb.create_sheet(sheet_name)

        # 表頭列
        for col, value in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=value)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER_THIN

        # 資料列
        for row_idx, row in enumerate(rows, 2):
            bg, fg = _MANTIS_COLORS.get(row.get("category", "normal"), ("111827", "E2E8F0"))
            fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            font = Font(name="微軟正黑體", size=10, color=fg)
            values = [
                row_idx - 1,
                row["ticket_id"],
                row["summary"],
                row["status"],
                row["priority"],
                row["unresolved_days"],
                row["last_updated"],
                row["handler"],
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = fill
                cell.font = font
                cell.border = BORDER_THIN

        # 凍結首列
        ws.freeze_panes = "A2"

        # 自動調整欄寬
        for col_cells in ws.columns:
            max_length = max(
                (len(str(c.value)) for c in col_cells if c.value), default=0
            )
            adjusted = min(max_length + 2, 50)
            if adjusted > 0:
                ws.column_dimensions[col_cells[0].column_letter].width = adjusted

        wb.save(str(path))

    @staticmethod
    def append_stats_chart_sheet(path: Path, stats: dict, start: str, end: str) -> None:
        """將統計圖表（matplotlib PNG）嵌入 Excel 第一個工作表位置。"""
        import io

        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.font_manager as fm
        import matplotlib.pyplot as plt
        from openpyxl.drawing.image import Image as XLImage

        _CJK = ["Microsoft JhengHei", "Microsoft YaHei", "SimHei", "PingFang TC"]
        _avail = {f.name for f in fm.fontManager.ttflist}
        _font = next((f for f in _CJK if f in _avail), None)
        if _font:
            plt.rcParams["font.family"] = _font
        plt.rcParams["axes.unicode_minus"] = False

        BG = "#FFFFFF"
        COLORS = ["#1E3A5F", "#10b981", "#f59e0b", "#8b5cf6", "#ef4444", "#06b6d4"]

        cs_stats = stats["cs_stats"]
        issue_counts = stats["issue_type_counts"]
        monthly = stats["monthly_counts"]
        customer_counts = stats.get("customer_counts", {})
        reply_dist = stats.get("reply_dist", {})

        names = [s["name"] for s in cs_stats]
        totals = [s["total"] for s in cs_stats]
        actives = [s["active"] for s in cs_stats]
        dones = [s["done"] for s in cs_stats]
        companies_cnt = [s["companies"] for s in cs_stats]

        fig = plt.figure(figsize=(16, 18), facecolor=BG)
        fig.suptitle(f"客服統計儀表板  {start} ～ {end}", fontsize=14, fontweight="bold", y=0.99)

        def _style_xl(ax):
            ax.grid(axis="y", linestyle="--", alpha=0.4)

        ax1 = fig.add_subplot(3, 2, 1)
        x = range(len(names))
        w = 0.28
        ax1.bar([i - w for i in x], totals, width=w, label="總案件", color=COLORS[0])
        ax1.bar(list(x), actives, width=w, label="處理中", color=COLORS[2])
        ax1.bar([i + w for i in x], dones, width=w, label="已完成", color=COLORS[1])
        ax1.set_xticks(list(x))
        ax1.set_xticklabels(names)
        ax1.set_title("各客服案件數", fontsize=11)
        ax1.legend(fontsize=8)
        _style_xl(ax1)
        for i, v in enumerate(totals):
            ax1.text(i - w, v + 0.1, str(v), ha="center", fontsize=9)

        ax2 = fig.add_subplot(3, 2, 2)
        bars = ax2.bar(names, companies_cnt, color=COLORS[:len(names)])
        ax2.set_title("各客服負責客戶數", fontsize=11)
        _style_xl(ax2)
        for bar, v in zip(bars, companies_cnt):
            ax2.text(bar.get_x() + bar.get_width() / 2, v + 0.1, str(v), ha="center", fontsize=10)

        ax3 = fig.add_subplot(3, 2, 3)
        if issue_counts:
            sorted_issues = sorted(issue_counts.items(), key=lambda x: -x[1])
            top_n = sorted_issues[:6]
            other_sum = sum(v for _, v in sorted_issues[6:])
            if other_sum:
                top_n.append(("其他", other_sum))
            labels = [k for k, _ in top_n]
            sizes = [v for _, v in top_n]
            ax3.pie(sizes, labels=labels, autopct="%1.1f%%",
                    colors=COLORS[:len(labels)], startangle=140, textprops={"fontsize": 9})
        ax3.set_title("問題類型分佈", fontsize=11)

        ax4 = fig.add_subplot(3, 2, 4)
        if monthly:
            months = list(monthly.keys())
            counts = list(monthly.values())
            ax4.plot(months, counts, color=COLORS[0], marker="o", linewidth=2, markersize=6)
            ax4.fill_between(months, counts, alpha=0.1, color=COLORS[0])
            ax4.set_xticks(range(len(months)))
            ax4.set_xticklabels(months, rotation=45, ha="right", fontsize=8)
            for i, v in enumerate(counts):
                ax4.text(i, v + 0.2, str(v), ha="center", fontsize=8)
        ax4.set_title("月份案件量趨勢", fontsize=11)
        ax4.grid(linestyle="--", alpha=0.4)

        ax5 = fig.add_subplot(3, 2, 5)
        ax5.grid(axis="x", linestyle="--", alpha=0.4)
        if customer_counts:
            cnames = list(customer_counts.keys())
            ccounts = list(customer_counts.values())
            y_pos = range(len(cnames))
            ax5.barh(list(y_pos), ccounts, color=COLORS[0])
            ax5.set_yticks(list(y_pos))
            ax5.set_yticklabels(cnames, fontsize=8)
            for i, v in enumerate(ccounts):
                ax5.text(v + 0.1, i, str(v), va="center", fontsize=8)
        ax5.set_title("各客戶案件數 Top 15", fontsize=11)
        ax5.invert_yaxis()

        ax6 = fig.add_subplot(3, 2, 6)
        _style_xl(ax6)
        if reply_dist:
            r_labels = list(reply_dist.keys())
            r_vals = list(reply_dist.values())
            bars6 = ax6.bar(r_labels, r_vals, color=COLORS[3])
            for bar, v in zip(bars6, r_vals):
                ax6.text(bar.get_x() + bar.get_width() / 2, v + 0.1, str(v),
                         ha="center", fontsize=9)
        ax6.set_title("案件回覆次數分佈", fontsize=11)

        fig.tight_layout(rect=[0, 0, 1, 0.98])

        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, facecolor=BG, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)

        wb = openpyxl.load_workbook(str(path))
        ws = wb.create_sheet("📊 統計圖表", 0)
        img = XLImage(buf)
        img.anchor = "A1"
        ws.add_image(img)
        wb.save(str(path))
