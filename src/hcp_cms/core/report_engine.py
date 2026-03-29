"""Excel report generation — tracking table and monthly report."""

import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from hcp_cms.data.models import Case
from hcp_cms.data.repositories import (
    CaseMantisRepository,
    CaseRepository,
    CompanyRepository,
    MantisRepository,
    QARepository,
)

# XML 1.0 不允許的控制字元（openpyxl IllegalCharacterError 的根源）
_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")

# Style constants
FONT_HEADER = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
FILL_ALT_ROW = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _clean(v: Any) -> Any:
    """過濾 XML 1.0 非法控制字元，避免 openpyxl IllegalCharacterError。"""
    if isinstance(v, str):
        return _ILLEGAL_CHARS_RE.sub("", v)
    return v


def _clean_row(row: list[Any]) -> list[Any]:
    """對整列每個值套用 _clean()。"""
    return [_clean(v) for v in row]


def _reply_hours(sent: str | None, replied: str | None) -> str:
    """計算首次回覆時效（小時），無法計算時回傳空字串。"""
    if not sent or not replied:
        return ""
    fmt = "%Y/%m/%d %H:%M:%S"
    try:
        delta = datetime.strptime(replied, fmt) - datetime.strptime(sent, fmt)
        hours = delta.total_seconds() / 3600
        return f"{hours:.1f}"
    except ValueError:
        return ""


class ReportEngine:
    def __init__(self, conn: sqlite3.Connection) -> None:
        self._conn = conn
        self._case_repo = CaseRepository(conn)
        self._qa_repo = QARepository(conn)
        self._mantis_repo = MantisRepository(conn)
        self._company_repo = CompanyRepository(conn)
        self._case_mantis_repo = CaseMantisRepository(conn)
        from hcp_cms.core.custom_column_manager import CustomColumnManager
        self._custom_col_mgr = CustomColumnManager(conn)

    def generate_tracking_table(self, start_date: str, end_date: str, output_path: Path) -> Path:
        """Generate tracking table Excel with multiple sheets.

        Args:
            start_date: 起始日期，格式 YYYY/MM/DD
            end_date:   結束日期，格式 YYYY/MM/DD（含當天）
        """
        cases = self._case_repo.list_by_date_range(start_date, end_date)
        qas = self._qa_repo.list_all()
        companies = self._company_repo.list_all()
        mantis_tickets = self._mantis_repo.list_all()

        # 建立公司 id → Company 快取
        company_map = {c.company_id: c for c in companies}

        wb = openpyxl.Workbook()

        # 預先計算各公司頁籤名稱（供快速連結使用）
        company_cases: dict[str, list[Case]] = {}
        for case in cases:
            cid = case.company_id or "unknown"
            company_cases.setdefault(cid, []).append(case)

        # Excel 工作表名稱不可含 \ / * ? : [ ]，一律替換為 -
        _INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')

        def _safe_sheet_name(raw: str, suffix: str = "_問題") -> str:
            sanitized = _INVALID_SHEET_CHARS.sub("-", raw)
            return (sanitized[:28] + suffix)[:31]

        company_sheet_names: dict[str, str] = {}
        for comp in companies:
            if company_cases.get(comp.company_id):
                raw = f"{comp.domain}({comp.name})" if comp.domain else comp.name
                company_sheet_names[comp.company_id] = _safe_sheet_name(raw)

        # ── Sheet 1: 客戶索引 ──────────────────────────────────────────
        ws = wb.active
        ws.title = "📋 客戶索引"
        self._write_header(ws, ["#", "公司名稱", "Email 域名", "聯絡方式", "案件數", "快速連結"])
        for i, comp in enumerate(companies, 1):
            count = sum(1 for c in cases if c.company_id == comp.company_id)
            ws.append(_clean_row([i, comp.name, comp.domain, comp.contact_info or "", count, ""]))
            self._style_data_row(ws, i + 1)
            if comp.company_id in company_sheet_names:
                sn = company_sheet_names[comp.company_id]
                lc = ws.cell(row=i + 1, column=6)
                lc.value = f'=HYPERLINK("#\'{sn}\'!A1","→ {comp.name}問題記錄")'
                lc.font = Font(name="微軟正黑體", size=11, color="0563C1", underline="single")

        # ── Sheet 2: 問題追蹤總表 ──────────────────────────────────────
        ws2 = wb.create_sheet("問題追蹤總表")
        custom_cols = self._custom_col_mgr.list_columns()
        main_headers = [
            "案件編號", "聯絡方式", "問題狀態", "優先等級",
            "寄件時間", "首次回覆時效(hr)", "客戶", "客戶公司", "客戶聯絡電話",
            "主旨", "系統／產品", "問題類型", "錯誤類型",
            "受影響員工人數", "影響期間", "處理進度", "負責人",
            "預計回覆時間", "實際回覆時間", "結案時間",
            "是否需升級", "是否有附圖", "QA文件名稱", "備註",
        ]
        main_headers = main_headers + [col.col_label for col in custom_cols]
        self._write_header(ws2, main_headers)
        for i, case in enumerate(cases, 1):
            comp = company_map.get(case.company_id or "")
            company_name = comp.name if comp else (case.company_id or "")
            company_phone = comp.contact_info if comp else ""
            closed_at = case.updated_at if case.status in ("已完成", "Closed") else ""
            ws2.append(_clean_row([
                case.case_id, case.contact_method, case.status, case.priority,
                case.sent_time, _reply_hours(case.sent_time, case.actual_reply),
                case.contact_person, company_name, company_phone or "",
                case.subject, case.system_product, case.issue_type, case.error_type,
                "", case.impact_period, case.progress, case.handler,
                "", case.actual_reply, closed_at,
                "", "", "", case.notes or "",
            ] + [_clean(case.extra_fields.get(col.col_key, "")) for col in custom_cols]))
            self._style_data_row(ws2, i + 1)

        # ── Sheet 3: QA知識庫 ──────────────────────────────────────────
        ws3 = wb.create_sheet("QA知識庫")
        qa_headers = [
            "QA編號", "系統／產品", "問題類型", "錯誤類型",
            "Q｜客戶問題描述", "A｜標準回覆內容",
            "附圖說明", "Word文件名稱", "建立日期", "建立人", "備註",
        ]
        self._write_header(ws3, qa_headers)
        for i, qa in enumerate(qas, 1):
            ws3.append(_clean_row([
                qa.qa_id, qa.system_product, qa.issue_type, qa.error_type,
                qa.question, qa.answer,
                qa.has_image, qa.doc_name or "", qa.created_at or "",
                qa.created_by or "", qa.notes or "",
            ]))
            self._style_data_row(ws3, i + 1)

        # ── 個別公司頁籤 ───────────────────────────────────────────────
        comp_case_headers = [
            "案件編號", "聯絡方式", "問題狀態", "優先等級",
            "寄件時間", "主旨", "系統／產品", "問題類型", "錯誤類型",
            "影響期間", "處理進度", "實際回覆時間", "結案時間", "備註",
        ] + [col.col_label for col in custom_cols]

        for comp in companies:
            comp_cases = company_cases.get(comp.company_id, [])
            if not comp_cases:
                continue
            sheet_name = company_sheet_names[comp.company_id]
            ws_c = wb.create_sheet(sheet_name)
            # Row 1：返回客戶索引連結
            back_cell = ws_c.cell(row=1, column=1,
                value='=HYPERLINK("#\'📋 客戶索引\'!A1","↩ 返回客戶索引")')
            back_cell.font = Font(name="微軟正黑體", size=10, color="0563C1", underline="single")
            # Row 2：表頭
            self._write_header(ws_c, comp_case_headers, row_num=2)
            # Row 3+：資料
            for i, case in enumerate(comp_cases, 1):
                closed_at = case.updated_at if case.status in ("已完成", "Closed") else ""
                row_num = i + 2
                ws_c.cell(row=row_num, column=1, value=_clean(case.case_id))
                ws_c.cell(row=row_num, column=2, value=_clean(case.contact_method))
                ws_c.cell(row=row_num, column=3, value=_clean(case.status))
                ws_c.cell(row=row_num, column=4, value=_clean(case.priority))
                ws_c.cell(row=row_num, column=5, value=_clean(case.sent_time))
                ws_c.cell(row=row_num, column=6, value=_clean(case.subject))
                ws_c.cell(row=row_num, column=7, value=_clean(case.system_product))
                ws_c.cell(row=row_num, column=8, value=_clean(case.issue_type))
                ws_c.cell(row=row_num, column=9, value=_clean(case.error_type))
                ws_c.cell(row=row_num, column=10, value=_clean(case.impact_period))
                ws_c.cell(row=row_num, column=11, value=_clean(case.progress))
                ws_c.cell(row=row_num, column=12, value=_clean(case.actual_reply))
                ws_c.cell(row=row_num, column=13, value=_clean(closed_at))
                ws_c.cell(row=row_num, column=14, value=_clean(case.notes or ""))
                for j, col in enumerate(custom_cols):
                    ws_c.cell(
                        row=row_num,
                        column=15 + j,
                        value=_clean(case.extra_fields.get(col.col_key, "")),
                    )
                self._style_data_row(ws_c, row_num)

        # ── Mantis提單追蹤 ─────────────────────────────────────────────
        ws_m = wb.create_sheet("Mantis提單追蹤")
        mantis_headers = [
            "Mantis票號", "建立時間", "客戶", "問題摘要", "關聯CS案件",
            "優先等級", "狀態", "類型", "相關程式/模組",
            "內部確認進度", "負責人", "預計修復日期", "實際修復日期", "備註",
        ]
        self._write_header(ws_m, mantis_headers)
        for i, ticket in enumerate(mantis_tickets, 1):
            comp = company_map.get(ticket.company_id or "")
            company_name = comp.name if comp else (ticket.company_id or "")
            linked = ", ".join(self._case_mantis_repo.get_cases_for_ticket(ticket.ticket_id))
            ws_m.append(_clean_row([
                ticket.ticket_id, ticket.created_time or "", company_name,
                ticket.summary, linked,
                ticket.priority or "", ticket.status or "", ticket.issue_type or "",
                ticket.module or "", ticket.progress or "", ticket.handler or "",
                ticket.planned_fix or "", ticket.actual_fix or "", ticket.notes or "",
            ]))
            self._style_data_row(ws_m, i + 1)

        # ── 客制需求 ───────────────────────────────────────────────────
        custom_cases = [c for c in cases if c.issue_type and "客制" in c.issue_type]
        if custom_cases:
            ws_custom = wb.create_sheet("客制需求")
            self._write_header(ws_custom, comp_case_headers)
            for i, case in enumerate(custom_cases, 1):
                closed_at = case.updated_at if case.status in ("已完成", "Closed") else ""
                ws_custom.append(_clean_row([
                    case.case_id, case.contact_method, case.status, case.priority,
                    case.sent_time, case.subject, case.system_product,
                    case.issue_type, case.error_type,
                    case.impact_period, case.progress, case.actual_reply,
                    closed_at, case.notes or "",
                ]))
                self._style_data_row(ws_custom, i + 1)

        wb.save(str(output_path))
        return output_path

    def build_tracking_table(self, start_date: str, end_date: str) -> dict[str, list[list]]:
        """組裝問題追蹤總表的所有工作表資料，回傳純資料結構（無 Excel 依賴）。

        Args:
            start_date: 起始日期，格式 YYYY/MM/DD
            end_date:   結束日期，格式 YYYY/MM/DD（含當天）

        Returns:
            dict，key 為工作表名稱，value 為 list of rows（第 0 列為表頭）。
        """
        cases = self._case_repo.list_by_date_range(start_date, end_date)
        qas = self._qa_repo.list_all()
        companies = self._company_repo.list_all()
        mantis_tickets = self._mantis_repo.list_all()

        # 建立公司 id → Company 快取
        company_map = {c.company_id: c for c in companies}

        # 預先計算各公司頁籤名稱
        company_cases: dict[str, list] = {}
        for case in cases:
            cid = case.company_id or "unknown"
            company_cases.setdefault(cid, []).append(case)

        # 工作表名稱不可含 \ / * ? : [ ]，一律替換為 -
        _INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')

        def _safe_sheet_name(raw: str, suffix: str = "_問題") -> str:
            sanitized = _INVALID_SHEET_CHARS.sub("-", raw)
            return (sanitized[:28] + suffix)[:31]

        company_sheet_names: dict[str, str] = {}
        for comp in companies:
            if company_cases.get(comp.company_id):
                raw = f"{comp.domain}({comp.name})" if comp.domain else comp.name
                company_sheet_names[comp.company_id] = _safe_sheet_name(raw)

        result: dict[str, list[list]] = {}

        # ── Sheet 1: 客戶索引 ──────────────────────────────────────────
        index_rows: list[list] = [["#", "公司名稱", "Email 域名", "聯絡方式", "案件數", "快速連結"]]
        for i, comp in enumerate(companies, 1):
            count = sum(1 for c in cases if c.company_id == comp.company_id)
            if comp.company_id in company_sheet_names:
                link_text = f"→ {comp.name}問題記錄"
            else:
                link_text = ""
            index_rows.append(_clean_row([i, comp.name, comp.domain, comp.contact_info or "", count, link_text]))
        result["📋 客戶索引"] = index_rows

        # ── Sheet 2: 問題追蹤總表 ──────────────────────────────────────
        custom_cols = self._custom_col_mgr.list_columns()
        main_headers = [
            "案件編號", "聯絡方式", "問題狀態", "優先等級",
            "寄件時間", "首次回覆時效(hr)", "客戶", "客戶公司", "客戶聯絡電話",
            "主旨", "系統／產品", "問題類型", "錯誤類型",
            "受影響員工人數", "影響期間", "處理進度", "負責人",
            "預計回覆時間", "實際回覆時間", "結案時間",
            "是否需升級", "是否有附圖", "QA文件名稱", "備註",
        ] + [col.col_label for col in custom_cols]

        tracking_rows: list[list] = [main_headers]
        for case in cases:
            comp = company_map.get(case.company_id or "")
            company_name = comp.name if comp else (case.company_id or "")
            company_phone = comp.contact_info if comp else ""
            closed_at = case.updated_at if case.status in ("已完成", "Closed") else ""
            tracking_rows.append(_clean_row([
                case.case_id, case.contact_method, case.status, case.priority,
                case.sent_time, _reply_hours(case.sent_time, case.actual_reply),
                case.contact_person, company_name, company_phone or "",
                case.subject, case.system_product, case.issue_type, case.error_type,
                "", case.impact_period, case.progress, case.handler,
                "", case.actual_reply, closed_at,
                "", "", "", case.notes or "",
            ] + [_clean(case.extra_fields.get(col.col_key, "")) for col in custom_cols]))
        result["問題追蹤總表"] = tracking_rows

        # ── Sheet 3: QA知識庫 ──────────────────────────────────────────
        qa_headers = [
            "QA編號", "系統／產品", "問題類型", "錯誤類型",
            "Q｜客戶問題描述", "A｜標準回覆內容",
            "附圖說明", "Word文件名稱", "建立日期", "建立人", "備註",
        ]
        qa_rows: list[list] = [qa_headers]
        for qa in qas:
            qa_rows.append(_clean_row([
                qa.qa_id, qa.system_product, qa.issue_type, qa.error_type,
                qa.question, qa.answer,
                qa.has_image, qa.doc_name or "", qa.created_at or "",
                qa.created_by or "", qa.notes or "",
            ]))
        result["QA知識庫"] = qa_rows

        # ── 個別公司頁籤 ───────────────────────────────────────────────
        comp_case_headers = [
            "案件編號", "聯絡方式", "問題狀態", "優先等級",
            "寄件時間", "主旨", "系統／產品", "問題類型", "錯誤類型",
            "影響期間", "處理進度", "實際回覆時間", "結案時間", "備註",
        ] + [col.col_label for col in custom_cols]

        for comp in companies:
            comp_cases = company_cases.get(comp.company_id, [])
            if not comp_cases:
                continue
            sheet_name = company_sheet_names[comp.company_id]
            # Row 0：返回客戶索引（純文字）
            # Row 1：表頭
            # Row 2+：資料
            comp_rows: list[list] = [
                ["↩ 返回客戶索引"],
                comp_case_headers,
            ]
            for case in comp_cases:
                closed_at = case.updated_at if case.status in ("已完成", "Closed") else ""
                comp_rows.append(_clean_row([
                    case.case_id, case.contact_method, case.status, case.priority,
                    case.sent_time, case.subject, case.system_product,
                    case.issue_type, case.error_type,
                    case.impact_period, case.progress, case.actual_reply,
                    closed_at, case.notes or "",
                ] + [_clean(case.extra_fields.get(col.col_key, "")) for col in custom_cols]))
            result[sheet_name] = comp_rows

        # ── Mantis提單追蹤 ─────────────────────────────────────────────
        mantis_headers = [
            "Mantis票號", "建立時間", "客戶", "問題摘要", "關聯CS案件",
            "優先等級", "狀態", "類型", "相關程式/模組",
            "內部確認進度", "負責人", "預計修復日期", "實際修復日期", "備註",
        ]
        mantis_rows: list[list] = [mantis_headers]
        for ticket in mantis_tickets:
            comp = company_map.get(ticket.company_id or "")
            company_name = comp.name if comp else (ticket.company_id or "")
            linked = ", ".join(self._case_mantis_repo.get_cases_for_ticket(ticket.ticket_id))
            mantis_rows.append(_clean_row([
                ticket.ticket_id, ticket.created_time or "", company_name,
                ticket.summary, linked,
                ticket.priority or "", ticket.status or "", ticket.issue_type or "",
                ticket.module or "", ticket.progress or "", ticket.handler or "",
                ticket.planned_fix or "", ticket.actual_fix or "", ticket.notes or "",
            ]))
        result["Mantis提單追蹤"] = mantis_rows

        # ── 客制需求 ───────────────────────────────────────────────────
        custom_cases = [c for c in cases if c.issue_type and "客制" in c.issue_type]
        if custom_cases:
            custom_rows: list[list] = [comp_case_headers]
            for case in custom_cases:
                closed_at = case.updated_at if case.status in ("已完成", "Closed") else ""
                custom_rows.append(_clean_row([
                    case.case_id, case.contact_method, case.status, case.priority,
                    case.sent_time, case.subject, case.system_product,
                    case.issue_type, case.error_type,
                    case.impact_period, case.progress, case.actual_reply,
                    closed_at, case.notes or "",
                ] + [_clean(case.extra_fields.get(col.col_key, "")) for col in custom_cols]))
            result["客制需求"] = custom_rows

        return result

    def build_monthly_report(self, start_date: str, end_date: str) -> dict[str, list[list]]:
        """組裝月報的所有工作表資料，回傳純資料結構（無 Excel 依賴）。

        Args:
            start_date: 起始日期，格式 YYYY/MM/DD
            end_date:   結束日期，格式 YYYY/MM/DD（含當天）

        Returns:
            dict，key 為工作表名稱，value 為 list of rows。
            - "📊 月報摘要": row[0] 標題列、row[1] 表頭、row[2:5] KPI、row[6] 空行、row[7] 類型標題、row[8] 類型表頭、row[9+] 類型統計
            - "📋 案件明細": row[0] 表頭、row[1:] 案件資料
            - "🏢 客戶分析": row[0] 表頭、row[1:] 各公司統計
        """
        cases = self._case_repo.list_by_date_range(start_date, end_date)
        companies = self._company_repo.list_all()
        company_map = {c.company_id: c for c in companies}

        closed_statuses = {"已完成", "Closed", "已回覆"}
        total = len(cases)
        replied = sum(1 for c in cases if c.status == "已回覆")
        pending = sum(1 for c in cases if c.status not in closed_statuses)
        reply_rate = (replied / total * 100) if total > 0 else 0.0

        result: dict[str, list[list]] = {}

        # ── Sheet 1: 月報摘要 ──────────────────────────────────────────
        summary_rows: list[list] = [
            # row 0: 標題列
            [f"📊 客服報表摘要 — {start_date} ～ {end_date}", f"產生日期：{datetime.now().strftime('%Y/%m/%d %H:%M')}"],
            # row 1: 表頭
            ["指標", "數值", "說明"],
            # row 2-5: KPI
            ["案件總數", total, f"{start_date} ～ {end_date}"],
            ["已回覆", replied, "replied = 是"],
            ["待處理", pending, "狀態非已完成/Closed"],
            ["回覆率", f"{reply_rate:.1f}%", "已回覆 ÷ 總數 × 100%"],
            # row 6: 空行
            [],
            # row 7: 問題類型統計標題
            ["問題類型統計"],
            # row 8: 類型表頭
            ["問題類型", "件數", "佔比"],
        ]
        # row 9+: 問題類型明細
        issue_counts: dict[str, int] = {}
        for c in cases:
            it = c.issue_type or "其他"
            issue_counts[it] = issue_counts.get(it, 0) + 1
        for it, count in sorted(issue_counts.items(), key=lambda x: -x[1]):
            pct = count / total * 100 if total > 0 else 0
            summary_rows.append([it, count, f"{pct:.1f}%"])
        result["📊 月報摘要"] = summary_rows

        # ── Sheet 2: 案件明細 ──────────────────────────────────────────
        detail_headers = [
            "案件編號", "聯絡方式", "狀態", "優先", "寄送時間",
            "客戶", "聯絡人", "主旨", "系統/產品", "問題類型", "錯誤類型",
            "影響期間", "進度", "實際回覆時間", "備註",
            "RD 負責人", "處理人", "回覆次數", "關聯案件",
        ]
        detail_rows: list[list] = [detail_headers]
        for case in cases:
            comp = company_map.get(case.company_id or "")
            company_name = comp.name if comp else (case.company_id or "")
            detail_rows.append(_clean_row([
                case.case_id, case.contact_method, case.status, case.priority,
                case.sent_time,
                company_name, case.contact_person, case.subject,
                case.system_product, case.issue_type, case.error_type,
                case.impact_period, case.progress, case.actual_reply,
                case.notes or "", case.rd_assignee or "", case.handler or "",
                case.reply_count, case.linked_case_id or "",
            ]))
        result["📋 案件明細"] = detail_rows

        # ── Sheet 3: 客戶分析 ──────────────────────────────────────────
        analysis_rows: list[list] = [["客戶", "已回覆", "處理中", "合計"]]
        company_stats: dict[str, dict[str, int]] = {}
        for case in cases:
            comp = company_map.get(case.company_id or "")
            cname = comp.name if comp else (case.company_id or "（未知）")
            if cname not in company_stats:
                company_stats[cname] = {"replied": 0, "pending": 0}
            if case.status == "已回覆":
                company_stats[cname]["replied"] += 1
            if case.status not in closed_statuses:
                company_stats[cname]["pending"] += 1
        for cname, stat in sorted(company_stats.items()):
            total_c = stat["replied"] + stat["pending"]
            analysis_rows.append(_clean_row([cname, stat["replied"], stat["pending"], total_c]))
        result["🏢 客戶分析"] = analysis_rows

        return result

    def generate_monthly_report(self, start_date: str, end_date: str, output_path: Path) -> Path:
        """Generate monthly report Excel with KPI summary.

        Args:
            start_date: 起始日期，格式 YYYY/MM/DD
            end_date:   結束日期，格式 YYYY/MM/DD（含當天）
        """
        cases = self._case_repo.list_by_date_range(start_date, end_date)
        companies = self._company_repo.list_all()
        company_map = {c.company_id: c for c in companies}

        wb = openpyxl.Workbook()

        closed_statuses = {"已完成", "Closed", "已回覆"}
        total = len(cases)
        replied = sum(1 for c in cases if c.status == "已回覆")
        pending = sum(1 for c in cases if c.status not in closed_statuses)
        reply_rate = (replied / total * 100) if total > 0 else 0.0

        # ── Sheet 1: 月報摘要 ──────────────────────────────────────────
        ws = wb.active
        ws.title = "📊 月報摘要"

        # 標題列
        title_cell = ws.cell(row=1, column=1, value=f"📊 客服報表摘要 — {start_date} ～ {end_date}")
        title_cell.font = Font(name="微軟正黑體", size=14, bold=True, color="1E3A5F")
        ws.cell(row=1, column=2, value=f"產生日期：{datetime.now().strftime('%Y/%m/%d %H:%M')}")

        # KPI rows
        self._write_header(ws, ["指標", "數值", "說明"], row_num=3)
        ws.append(["案件總數", total, f"{start_date} ～ {end_date}"])
        ws.append(["已回覆", replied, "replied = 是"])
        ws.append(["待處理", pending, "狀態非已完成/Closed"])
        ws.append(["回覆率", f"{reply_rate:.1f}%", "已回覆 ÷ 總數 × 100%"])

        # 問題類型統計
        ws.append([])
        type_title = ws.cell(row=ws.max_row + 1, column=1, value="問題類型統計")
        type_title.font = Font(name="微軟正黑體", size=11, bold=True)
        self._write_header(ws, ["問題類型", "件數", "佔比"], row_num=ws.max_row + 1)
        issue_counts: dict[str, int] = {}
        for c in cases:
            it = c.issue_type or "其他"
            issue_counts[it] = issue_counts.get(it, 0) + 1
        for it, count in sorted(issue_counts.items(), key=lambda x: -x[1]):
            pct = count / total * 100 if total > 0 else 0
            ws.append([it, count, f"{pct:.1f}%"])

        # ── Sheet 2: 案件明細 ──────────────────────────────────────────
        ws2 = wb.create_sheet("📋 案件明細")
        detail_headers = [
            "案件編號", "聯絡方式", "狀態", "優先", "寄送時間",
            "客戶", "聯絡人", "主旨", "系統/產品", "問題類型", "錯誤類型",
            "影響期間", "進度", "實際回覆時間", "備註",
            "RD 負責人", "處理人", "回覆次數", "關聯案件",
        ]
        self._write_header(ws2, detail_headers)
        for i, case in enumerate(cases, 1):
            comp = company_map.get(case.company_id or "")
            company_name = comp.name if comp else (case.company_id or "")
            ws2.append(_clean_row([
                case.case_id, case.contact_method, case.status, case.priority,
                case.sent_time,
                company_name, case.contact_person, case.subject,
                case.system_product, case.issue_type, case.error_type,
                case.impact_period, case.progress, case.actual_reply,
                case.notes or "", case.rd_assignee or "", case.handler or "",
                case.reply_count, case.linked_case_id or "",
            ]))
            self._style_data_row(ws2, i + 1)

        # ── Sheet 3: 客戶分析 ──────────────────────────────────────────
        ws3 = wb.create_sheet("🏢 客戶分析")
        self._write_header(ws3, ["客戶", "已回覆", "處理中", "合計"])
        # 按公司彙整
        company_stats: dict[str, dict[str, int]] = {}
        for case in cases:
            comp = company_map.get(case.company_id or "")
            cname = comp.name if comp else (case.company_id or "（未知）")
            if cname not in company_stats:
                company_stats[cname] = {"replied": 0, "pending": 0}
            if case.status == "已回覆":
                company_stats[cname]["replied"] += 1
            if case.status not in closed_statuses:
                company_stats[cname]["pending"] += 1
        for i, (cname, stat) in enumerate(sorted(company_stats.items()), 1):
            total_c = stat["replied"] + stat["pending"]
            ws3.append(_clean_row([cname, stat["replied"], stat["pending"], total_c]))
            self._style_data_row(ws3, i + 1)

        wb.save(str(output_path))
        return output_path

    def _get_company_name(self, company_id: str | None) -> str:
        if not company_id:
            return ""
        company = self._company_repo.get_by_id(company_id)
        return company.name if company else company_id

    def _write_header(self, ws: Any, headers: list[str], row_num: int = 1) -> None:
        """Write styled header row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER_THIN

    def _style_data_row(self, ws: Any, row_num: int) -> None:
        """Apply alternating row style and borders."""
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = BORDER_THIN
            if row_num % 2 == 0:
                cell.fill = FILL_ALT_ROW
