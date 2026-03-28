"""CSV 精靈自訂欄位完整流程整合測試。"""
import csv
from pathlib import Path

import openpyxl
import pytest

from hcp_cms.core.csv_import_engine import ConflictStrategy, CsvImportEngine
from hcp_cms.core.custom_column_manager import CustomColumnManager
from hcp_cms.core.report_engine import ReportEngine
from hcp_cms.data.database import DatabaseManager
from hcp_cms.data.repositories import CaseRepository


@pytest.fixture
def db(tmp_path: Path) -> DatabaseManager:
    mgr = DatabaseManager(tmp_path / "csv_wizard_integration.db")
    mgr.initialize()
    yield mgr
    mgr.close()


def _write_csv(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


class TestFullFlow:
    def test_create_import_verify_report(self, db: DatabaseManager, tmp_path: Path) -> None:
        """建立自訂欄 → 匯入 → DB 驗證 → 報表驗證。"""
        conn = db.connection
        engine = CsvImportEngine(conn)

        # 1. 建立自訂欄
        cols = engine.create_custom_columns([("來源系統", "來源系統")])
        assert cols[0].col_key == "cx_1"

        # 2. 匯入 CSV
        csv_path = tmp_path / "data.csv"
        _write_csv(csv_path, [{
            "主旨": "整合測試案件", "寄件時間": "2026/03/26 10:00:00",
            "公司": "測試公司整合", "來源系統": "SAP",
        }])
        mapping = {
            "主旨": "subject", "寄件時間": "sent_time",
            "公司": "company_id", "來源系統": "cx_1",
        }
        result = engine.execute(csv_path, mapping, ConflictStrategy.SKIP)
        assert result.success == 1

        # 3. 驗證 Case.extra_fields
        case_repo = CaseRepository(conn)
        cases = case_repo.list_all()
        assert len(cases) == 1
        assert cases[0].extra_fields.get("cx_1") == "SAP"

        # 4. 驗證 CustomColumnManager 可取得欄位
        mgr = CustomColumnManager(conn)
        custom_cols = mgr.list_columns()
        assert any(c.col_label == "來源系統" for c in custom_cols)

        # 5. 驗證報表含自訂欄
        report_engine = ReportEngine(conn)
        out = tmp_path / "report.xlsx"
        report_engine.generate_tracking_table("2026/03/01", "2026/03/31", out)
        wb = openpyxl.load_workbook(out)
        ws2 = wb["問題追蹤總表"]
        headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
        assert "來源系統" in headers
