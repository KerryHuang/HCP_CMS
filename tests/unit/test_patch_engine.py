"""SinglePatchEngine 單元測試。"""
from pathlib import Path

import pytest

from hcp_cms.core.patch_engine import SinglePatchEngine
from hcp_cms.data.database import DatabaseManager


@pytest.fixture
def conn():
    db = DatabaseManager(":memory:")
    db.initialize()
    yield db.connection
    db.connection.close()


@pytest.fixture
def patch_dir(tmp_path: Path) -> Path:
    (tmp_path / "form").mkdir()
    (tmp_path / "sql").mkdir()
    (tmp_path / "muti").mkdir()
    (tmp_path / "form" / "PAYROLL.fmx").write_text("form")
    (tmp_path / "sql" / "update.sql").write_text("sql")
    (tmp_path / "setup.bat").write_text("bat")
    (tmp_path / "ReleaseNote.docx").write_bytes(b"")
    return tmp_path


class TestScanPatchDir:
    def test_finds_form_sql_muti_files(self, conn, patch_dir):
        eng = SinglePatchEngine(conn)
        result = eng.scan_patch_dir(str(patch_dir))
        assert "PAYROLL.fmx" in result["form_files"]
        assert "update.sql" in result["sql_files"]
        assert result["setup_bat"] is True
        assert result["missing"] == []

    def test_missing_muti_not_error(self, conn, patch_dir):
        import shutil
        shutil.rmtree(patch_dir / "muti")
        eng = SinglePatchEngine(conn)
        result = eng.scan_patch_dir(str(patch_dir))
        assert result["muti_files"] == []
        assert "muti/" not in result["missing"]

    def test_missing_form_reported(self, conn, patch_dir):
        import shutil
        shutil.rmtree(patch_dir / "form")
        eng = SinglePatchEngine(conn)
        result = eng.scan_patch_dir(str(patch_dir))
        assert "form/" in result["missing"]

    def test_detects_release_note(self, conn, patch_dir):
        eng = SinglePatchEngine(conn)
        result = eng.scan_patch_dir(str(patch_dir))
        assert result["release_note"] is not None
        assert "ReleaseNote" in result["release_note"]


class TestReadReleaseDoc:
    def test_parse_docx_returns_issues(self, conn, tmp_path):
        docx = pytest.importorskip("docx", reason="python-docx 未安裝，跳過")
        doc = docx.Document()
        doc.add_paragraph("Bug Fix")
        doc.add_paragraph("0015659  薪資計算錯誤修正")
        doc.add_paragraph("0015660  加班費計算異常")
        doc.add_paragraph("Enhancement")
        doc.add_paragraph("0015661  新增匯出功能")
        p = tmp_path / "ReleaseNote.docx"
        doc.save(str(p))

        eng = SinglePatchEngine(conn)
        issues = eng.read_release_doc(str(p))
        assert len(issues) == 3
        assert issues[0]["issue_no"] == "0015659"
        assert issues[0]["issue_type"] == "BugFix"
        assert issues[0]["description"] == "薪資計算錯誤修正"
        assert issues[0]["region"] == "共用"
        assert issues[2]["issue_type"] == "Enhancement"

    def test_missing_file_returns_empty(self, conn, tmp_path):
        eng = SinglePatchEngine(conn)
        result = eng.read_release_doc(str(tmp_path / "nonexist.docx"))
        assert result == []


class TestGenerateExcelReports:
    @pytest.fixture
    def engine_with_patch(self, conn):
        from hcp_cms.data.models import PatchIssue, PatchRecord
        from hcp_cms.data.repositories import PatchRepository
        repo = PatchRepository(conn)
        pid = repo.insert_patch(PatchRecord(type="single", patch_dir="C:/test"))
        repo.insert_issue(PatchIssue(patch_id=pid, issue_no="0015659",
                                     issue_type="BugFix", region="TW",
                                     description="薪資計算錯誤", sort_order=1))
        repo.insert_issue(PatchIssue(patch_id=pid, issue_no="0015660",
                                     issue_type="Enhancement", region="共用",
                                     description="新增匯出功能", sort_order=2))
        return SinglePatchEngine(conn), pid

    def test_generates_three_files(self, engine_with_patch, tmp_path):
        eng, pid = engine_with_patch
        paths = eng.generate_excel_reports(pid, output_dir=str(tmp_path))
        assert len(paths) == 3
        for p in paths:
            assert Path(p).exists()

    def test_issue_list_has_tracking_columns(self, engine_with_patch, tmp_path):
        import openpyxl
        eng, pid = engine_with_patch
        paths = eng.generate_excel_reports(pid, output_dir=str(tmp_path))
        issue_list = next(p for p in paths if "Issue清單整理" in p)
        wb = openpyxl.load_workbook(issue_list)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "客服驗證" in headers
        assert "客戶測試結果" in headers

    def test_release_notice_no_tracking_columns(self, engine_with_patch, tmp_path):
        import openpyxl
        eng, pid = engine_with_patch
        paths = eng.generate_excel_reports(pid, output_dir=str(tmp_path))
        notice = next(p for p in paths if "發行通知" in p)
        wb = openpyxl.load_workbook(notice)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "客服驗證" not in headers


class TestGenerateTestScripts:
    @pytest.fixture
    def engine_with_patch(self, conn):
        from hcp_cms.data.models import PatchIssue, PatchRecord
        from hcp_cms.data.repositories import PatchRepository
        repo = PatchRepository(conn)
        pid = repo.insert_patch(PatchRecord(type="single"))
        repo.insert_issue(PatchIssue(patch_id=pid, issue_no="0015659",
                                     description="測試說明", test_direction="測試步驟"))
        return SinglePatchEngine(conn), pid

    def test_generates_three_script_files(self, engine_with_patch, tmp_path):
        eng, pid = engine_with_patch
        paths = eng.generate_test_scripts(pid, output_dir=str(tmp_path))
        assert len(paths) == 3
        names = [Path(p).name for p in paths]
        assert any("客服版" in n for n in names)
        assert any("客戶版" in n for n in names)
        assert any("追蹤表" in n and n.endswith(".xlsx") for n in names)
        for p in paths:
            assert Path(p).exists()


class TestLoadFromArchive:
    def test_parse_version_tag_extracts_ip_pattern(self, conn):
        eng = SinglePatchEngine(conn)
        assert eng._parse_version_tag("IP_合併_20261101_HCP11G.7z") == "IP_合併_20261101"

    def test_parse_version_tag_fallback_to_stem(self, conn):
        eng = SinglePatchEngine(conn)
        assert eng._parse_version_tag("MyPatch.7z") == "MyPatch"

    def test_parse_version_tag_long_stem_truncated(self, conn):
        eng = SinglePatchEngine(conn)
        tag = eng._parse_version_tag("A" * 30 + ".7z")
        assert len(tag) <= 20

    def test_load_from_archive_returns_tuple(self, conn, tmp_path):
        from unittest.mock import MagicMock, patch
        archive = tmp_path / "IP_合併_20261101.7z"
        archive.write_bytes(b"fake")
        extract_dir = tmp_path / "out"

        mock_z = MagicMock()
        mock_z.__enter__ = lambda s: mock_z
        mock_z.__exit__ = MagicMock(return_value=False)

        with patch("py7zr.SevenZipFile", return_value=mock_z), \
             patch.object(SinglePatchEngine, "scan_patch_dir",
                          return_value={"release_note": None, "form_files": [],
                                        "sql_files": [], "muti_files": [],
                                        "setup_bat": False, "install_guide": None,
                                        "missing": []}):
            eng = SinglePatchEngine(conn)
            patch_id, version_tag, issue_count = eng.load_from_archive(
                str(archive), str(extract_dir)
            )
        assert version_tag == "IP_合併_20261101"
        assert isinstance(patch_id, int)
        assert issue_count == 0

    def test_load_from_archive_loads_issues_from_release_note(self, conn, tmp_path):
        from unittest.mock import MagicMock, patch
        archive = tmp_path / "IP_合併_20261201.7z"
        archive.write_bytes(b"fake")
        extract_dir = tmp_path / "out"
        fake_release = str(tmp_path / "ReleaseNote.docx")

        mock_z = MagicMock()
        mock_z.__enter__ = lambda s: mock_z
        mock_z.__exit__ = MagicMock(return_value=False)

        fake_scan = {"release_note": fake_release, "form_files": ["A.fmx"],
                     "sql_files": [], "muti_files": [],
                     "setup_bat": False, "install_guide": None, "missing": []}
        fake_issues = [{"issue_no": "0015659", "issue_type": "BugFix",
                        "description": "修正", "region": "TW"}]

        with patch("py7zr.SevenZipFile", return_value=mock_z), \
             patch.object(SinglePatchEngine, "scan_patch_dir", return_value=fake_scan), \
             patch.object(SinglePatchEngine, "read_release_doc", return_value=fake_issues):
            eng = SinglePatchEngine(conn)
            patch_id, version_tag, issue_count = eng.load_from_archive(
                str(archive), str(extract_dir)
            )
        assert issue_count == 1
        assert version_tag == "IP_合併_20261201"
