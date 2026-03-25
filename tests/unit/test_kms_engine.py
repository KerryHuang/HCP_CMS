"""Tests for KMSEngine."""

from pathlib import Path

import openpyxl
import pytest

from hcp_cms.core.kms_engine import KMSEngine
from hcp_cms.data.database import DatabaseManager
from hcp_cms.data.models import Case
from hcp_cms.data.repositories import CaseRepository


@pytest.fixture
def db(tmp_db_path: Path) -> DatabaseManager:
    db = DatabaseManager(tmp_db_path)
    db.initialize()
    yield db
    db.close()


@pytest.fixture
def kms(db: DatabaseManager) -> KMSEngine:
    return KMSEngine(db.connection)


class TestKMSEngine:
    def test_create_and_search(self, kms):
        kms.create_qa(question="薪資如何計算", answer="進入薪資模組")
        results = kms.search("薪資")
        assert len(results) > 0
        assert results[0].question == "薪資如何計算"

    def test_search_with_filter(self, kms):
        kms.create_qa(question="HCP 薪資", answer="A", system_product="HCP")
        kms.create_qa(question="ERP 薪資", answer="B", system_product="ERP")
        results = kms.search("薪資", system_product="HCP")
        assert all(r.system_product == "HCP" for r in results)

    def test_update_qa(self, kms):
        qa = kms.create_qa(question="舊問題", answer="舊回覆")
        updated = kms.update_qa(qa.qa_id, question="新問題", answer="新回覆")
        assert updated.question == "新問題"
        # Search should find updated content
        results = kms.search("新問題")
        assert len(results) > 0

    def test_delete_qa(self, kms):
        qa = kms.create_qa(question="待刪除", answer="test")
        kms.delete_qa(qa.qa_id)
        results = kms.search("待刪除")
        assert len(results) == 0

    def test_auto_extract_with_question(self, kms, db):
        case = Case(
            case_id="CS-2026-001",
            subject="請問薪資如何計算",
            progress="進入薪資模組設定",
            system_product="HCP",
        )
        CaseRepository(db.connection).insert(case)
        qa = kms.auto_extract_qa(case)
        assert qa is not None
        assert qa.source == "email"
        assert qa.source_case_id == "CS-2026-001"

    def test_auto_extract_no_question(self, kms):
        case = Case(case_id="CS-2026-002", subject="系統更新通知", progress="已更新")
        qa = kms.auto_extract_qa(case)
        assert qa is None

    def test_import_export_roundtrip(self, kms, tmp_path):
        # Create some QAs
        kms.create_qa(question="Q1", answer="A1", solution="S1", keywords="k1")
        kms.create_qa(question="Q2", answer="A2")

        # Export
        export_path = tmp_path / "qa_export.xlsx"
        kms.export_to_excel(export_path)
        assert export_path.exists()

        # Verify Excel content
        wb = openpyxl.load_workbook(str(export_path))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2
        wb.close()

    def test_import_from_excel(self, kms, tmp_path):
        # Create Excel file
        import_path = tmp_path / "qa_import.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["問題", "回覆", "解決方案", "關鍵字"])
        ws.append(["如何設定薪資", "進入設定頁面", "步驟1...", "薪資 設定"])
        ws.append(["請假規則", "參考手冊", None, None])
        wb.save(str(import_path))

        count = kms.import_from_excel(import_path)
        assert count == 2

        # Verify searchable
        results = kms.search("薪資")
        assert len(results) > 0


class TestKMSEngineStatus:
    def test_create_qa_待審核_不進_FTS(self, kms):
        qa = kms.create_qa(question="問題A", answer="回覆A", status="待審核")
        results = kms.search("問題A")
        assert all(r.qa_id != qa.qa_id for r in results)

    def test_create_qa_已完成_進_FTS(self, kms):
        kms.create_qa(question="問題B", answer="回覆B", status="已完成")
        results = kms.search("問題B")
        assert len(results) > 0

    def test_update_qa_待審核_不更新_FTS(self, kms):
        qa = kms.create_qa(question="問題C", answer="回覆C", status="待審核")
        kms.update_qa(qa.qa_id, answer="新回覆C")
        results = kms.search("問題C")
        assert all(r.qa_id != qa.qa_id for r in results)

    def test_update_qa_已完成_更新_FTS(self, kms):
        qa = kms.create_qa(question="問題D", answer="回覆D")
        kms.update_qa(qa.qa_id, answer="新回覆D")
        results = kms.search("問題D")
        assert len(results) > 0

    def test_update_qa_降級被拒絕(self, kms):
        qa = kms.create_qa(question="問題E", answer="回覆E", status="已完成")
        result = kms.update_qa(qa.qa_id, status="待審核")
        assert result is None
        from hcp_cms.data.repositories import QARepository
        assert QARepository(kms._conn).get_by_id(qa.qa_id).status == "已完成"

    def test_extract_qa_from_email_有_thread_question(self, kms, db):
        from hcp_cms.data.models import Case
        from hcp_cms.data.repositories import CaseRepository
        from hcp_cms.services.mail.base import RawEmail
        CaseRepository(db.connection).insert(Case(case_id="CS-001", subject="測試案件"))
        raw = RawEmail(thread_question="客戶詢問問題", thread_answer="我方回覆")
        qa = kms.extract_qa_from_email(raw, case_id="CS-001")
        assert qa is not None
        assert qa.status == "待審核"
        assert qa.question == "客戶詢問問題"
        assert qa.source_case_id == "CS-001"

    def test_extract_qa_from_email_無_thread_question(self, kms):
        from hcp_cms.services.mail.base import RawEmail
        raw = RawEmail(thread_question=None, thread_answer=None)
        qa = kms.extract_qa_from_email(raw)
        assert qa is None

    def test_approve_qa_更新_status_且_FTS_索引建立(self, kms):
        from hcp_cms.services.mail.base import RawEmail
        raw = RawEmail(thread_question="問題F", thread_answer="回覆F")
        qa = kms.extract_qa_from_email(raw)
        approved = kms.approve_qa(qa.qa_id, answer="修改後回覆F")
        assert approved is not None
        assert approved.status == "已完成"
        results = kms.search("問題F")
        assert len(results) > 0

    def test_list_pending(self, kms):
        kms.create_qa(question="待審1", answer="a", status="待審核")
        kms.create_qa(question="已完成1", answer="b", status="已完成")
        pending = kms.list_pending()
        assert len(pending) == 1
        assert pending[0].question == "待審1"

    def test_search_不返回待審核(self, kms):
        kms.create_qa(question="共同關鍵字", answer="a", status="待審核")
        kms.create_qa(question="共同關鍵字", answer="b", status="已完成")
        results = kms.search("共同關鍵字")
        assert all(r.status == "已完成" for r in results)

    def test_delete_待審核_QA_無害(self, kms):
        from hcp_cms.services.mail.base import RawEmail
        raw = RawEmail(thread_question="刪除測試", thread_answer="a")
        qa = kms.extract_qa_from_email(raw)
        kms.delete_qa(qa.qa_id)
        assert kms._qa_repo.get_by_id(qa.qa_id) is None

    def test_export_to_excel_排除待審核(self, kms, tmp_path):
        kms.create_qa(question="已完成QA", answer="a", status="已完成")
        kms.create_qa(question="待審核QA", answer="b", status="待審核")
        path = kms.export_to_excel(tmp_path / "out.xlsx")
        import openpyxl
        wb = openpyxl.load_workbook(str(path))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        questions = [r[1] for r in rows if r[1]]
        assert "已完成QA" in questions
        assert "待審核QA" not in questions

    # --- Task 4 tests ---

    def test_attach_images_nonexistent_msg(self, kms, tmp_path):
        """msg 不存在時回傳 0，DB 不更新。"""
        qa = kms.create_qa(question="測試", answer="答覆")
        count = kms.attach_images(qa.qa_id, tmp_path / "notexist.msg", tmp_path)
        assert count == 0
        qa_after = kms._qa_repo.get_by_id(qa.qa_id)
        assert qa_after.has_image == "否"  # 未更新

    def test_attach_images_updates_db(self, kms, tmp_path):
        """提供假 msg 路徑（存在但無圖片），has_image 更新為是。"""
        fake_msg = tmp_path / "fake.msg"
        fake_msg.write_bytes(b"")
        qa = kms.create_qa(question="圖片測試", answer="答覆")
        kms.attach_images(qa.qa_id, fake_msg, tmp_path)
        qa_after = kms._qa_repo.get_by_id(qa.qa_id)
        assert qa_after.has_image == "是"
        assert qa_after.doc_name == str(fake_msg)

    def test_extract_qa_from_email_with_db_dir(self, kms, db, tmp_path):
        """帶 db_dir 時，extract_qa_from_email 成功建立 QA 並記錄 doc_name。"""
        from hcp_cms.data.models import Case
        from hcp_cms.data.repositories import CaseRepository
        from hcp_cms.services.mail.base import RawEmail
        CaseRepository(db.connection).insert(Case(case_id="CS-001", subject="測試案件"))
        email = RawEmail(
            sender="customer@client.com",
            subject="測試",
            body="您好，請問如何操作？",
            thread_question="請問如何操作？",
            thread_answer="操作說明如下",
            source_file=str(tmp_path / "test.msg"),
        )
        # 建立假 msg 檔（讓 attach_images 不因不存在而略過）
        (tmp_path / "test.msg").write_bytes(b"")
        qa = kms.extract_qa_from_email(email, case_id="CS-001", db_dir=tmp_path)
        assert qa is not None
        assert qa.doc_name == str(tmp_path / "test.msg")

    # --- Task 5 tests ---

    def test_export_to_docx_creates_file(self, kms, tmp_path):
        # 必須建立「已完成」的 QA，qa_list=None 時呼叫 list_approved()
        kms.create_qa(question="問題一", answer="回覆一", status="已完成")
        kms.create_qa(question="問題二", answer="回覆二", status="已完成")
        out = tmp_path / "export.docx"
        result = kms.export_to_docx(out, db_dir=tmp_path)
        assert result == out
        assert out.exists()

    def test_export_to_docx_empty_list(self, kms, tmp_path):
        """空列表時建立含『無資料』標題的空白 docx，不拋例外。"""
        out = tmp_path / "empty.docx"
        kms.export_to_docx(out, db_dir=tmp_path, qa_list=[])
        assert out.exists()

    def test_export_to_docx_qa_list_override(self, kms, tmp_path):
        """傳入 qa_list 時只匯出指定的 QA。"""
        qa1 = kms.create_qa(question="匯出這筆", answer="A")
        kms.create_qa(question="不匯出這筆", answer="B")
        out = tmp_path / "partial.docx"
        kms.export_to_docx(out, db_dir=tmp_path, qa_list=[qa1])
        # 讀取確認只有 qa1
        from docx import Document
        doc = Document(out)
        full_text = " ".join(p.text for p in doc.paragraphs)
        assert "匯出這筆" in full_text
        assert "不匯出這筆" not in full_text
