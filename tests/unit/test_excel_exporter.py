"""Unit tests for ExcelExporter."""

from __future__ import annotations

import openpyxl

from hcp_cms.core.excel_exporter import ExcelExporter
from hcp_cms.core.sent_mail_manager import EnrichedSentMail


def _make_mail(
    date: str = "2026/03/27 10:00:00",
    recipients: list[str] | None = None,
    subject: str = "測試主旨",
    company_id: str | None = "C001",
    company_name: str | None = "測試公司",
    linked_case_id: str | None = None,
    company_reply_count: int = 3,
) -> EnrichedSentMail:
    return EnrichedSentMail(
        date=date,
        recipients=recipients or ["test@example.com"],
        subject=subject,
        company_id=company_id,
        company_name=company_name,
        linked_case_id=linked_case_id,
        company_reply_count=company_reply_count,
    )


class TestExcelExporter:
    def test_creates_file(self, tmp_path):
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail([_make_mail()], path)
        assert (tmp_path / "output.xlsx").exists()

    def test_sheet_names(self, tmp_path):
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail([_make_mail()], path)
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["公司彙總", "寄件清單"]

    def test_summary_header(self, tmp_path):
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail([_make_mail()], path)
        wb = openpyxl.load_workbook(path)
        ws = wb["公司彙總"]
        assert ws.cell(1, 1).value == "公司名稱"
        assert ws.cell(1, 2).value == "次數"

    def test_summary_rows_count(self, tmp_path):
        mails = [
            _make_mail(company_id="C001", company_name="甲公司", subject="主旨A"),
            _make_mail(company_id="C002", company_name="乙公司", subject="主旨B"),
            _make_mail(company_id="C001", company_name="甲公司", subject="主旨A"),  # 同主旨重複
        ]
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail(mails, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["公司彙總"]
        # 2 間公司 + 1 標題列
        assert ws.max_row == 3

    def test_summary_count_unique_subjects(self, tmp_path):
        """彙總次數 = 不重複主旨數（同主旨多封算一封）。"""
        mails = [
            _make_mail(company_id="C001", company_name="甲公司", subject="主旨A"),
            _make_mail(company_id="C001", company_name="甲公司", subject="主旨B"),
            _make_mail(company_id="C001", company_name="甲公司", subject="主旨A"),  # 重複，不計
        ]
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail(mails, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["公司彙總"]
        assert ws.cell(2, 2).value == 2  # 2 個不重複主旨

    def test_summary_sorted_by_count_desc(self, tmp_path):
        mails = [
            _make_mail(company_id="C001", company_name="甲公司", subject="主旨A"),
            _make_mail(company_id="C002", company_name="乙公司", subject="主旨B"),
            _make_mail(company_id="C002", company_name="乙公司", subject="主旨C"),
        ]
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail(mails, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["公司彙總"]
        assert ws.cell(2, 1).value == "乙公司"  # 2 主旨，排前面
        assert ws.cell(3, 1).value == "甲公司"  # 1 主旨

    def test_list_header(self, tmp_path):
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail([_make_mail()], path)
        wb = openpyxl.load_workbook(path)
        ws = wb["寄件清單"]
        headers = [ws.cell(1, c).value for c in range(1, 7)]
        assert headers == ["日期", "收件人", "主旨", "公司", "案件", "第幾封"]

    def test_list_rows_count(self, tmp_path):
        mails = [_make_mail(), _make_mail(), _make_mail()]
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail(mails, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["寄件清單"]
        # 3 封信 + 1 標題列
        assert ws.max_row == 4

    def test_list_recipients_joined(self, tmp_path):
        mail = _make_mail(recipients=["a@x.com", "b@x.com"])
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail([mail], path)
        wb = openpyxl.load_workbook(path)
        ws = wb["寄件清單"]
        assert ws.cell(2, 2).value == "a@x.com, b@x.com"

    def test_list_empty_case_filled_dash(self, tmp_path):
        mail = _make_mail(linked_case_id=None)
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail([mail], path)
        wb = openpyxl.load_workbook(path)
        ws = wb["寄件清單"]
        assert ws.cell(2, 5).value == "—"

    def test_list_empty_company_name_filled_dash(self, tmp_path):
        mail = _make_mail(company_id=None, company_name=None, company_reply_count=0)
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail([mail], path)
        wb = openpyxl.load_workbook(path)
        ws = wb["寄件清單"]
        assert ws.cell(2, 4).value == "—"

    def test_list_empty_company_id_no_count(self, tmp_path):
        mail = _make_mail(company_id=None, company_name=None, company_reply_count=0)
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail([mail], path)
        wb = openpyxl.load_workbook(path)
        ws = wb["寄件清單"]
        assert ws.cell(2, 6).value == "—"

    def test_list_full_datetime(self, tmp_path):
        """日期欄顯示完整日期時間，不截斷。"""
        mail = _make_mail(date="2026/03/27 14:30:00")
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail([mail], path)
        wb = openpyxl.load_workbook(path)
        ws = wb["寄件清單"]
        assert ws.cell(2, 1).value == "2026/03/27 14:30:00"

    def test_list_sequential_counter_per_subject(self, tmp_path):
        """同公司同主旨的信件依出現順序編號；不同主旨各自從 1 開始。"""
        mails = [
            _make_mail(company_id="C001", company_name="甲公司", subject="RE: 薪資問題"),
            _make_mail(company_id="C001", company_name="甲公司", subject="RE: 請假申請"),
            _make_mail(company_id="C001", company_name="甲公司", subject="RE: 薪資問題"),
            _make_mail(company_id="C002", company_name="乙公司", subject="RE: 薪資問題"),
        ]
        path = str(tmp_path / "output.xlsx")
        ExcelExporter().export_sent_mail(mails, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["寄件清單"]
        assert ws.cell(2, 6).value == "1"  # 甲公司「薪資問題」第1封
        assert ws.cell(3, 6).value == "1"  # 甲公司「請假申請」第1封（不同主旨重新計）
        assert ws.cell(4, 6).value == "2"  # 甲公司「薪資問題」第2封
        assert ws.cell(5, 6).value == "1"  # 乙公司「薪資問題」第1封（不同公司重新計）
