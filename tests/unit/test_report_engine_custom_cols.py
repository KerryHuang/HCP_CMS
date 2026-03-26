"""ReportEngine 自訂欄位整合測試。"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from hcp_cms.core.report_engine import ReportEngine
from hcp_cms.data.database import DatabaseManager
from hcp_cms.data.repositories import CustomColumnRepository


@pytest.fixture
def db(tmp_db_path: Path) -> DatabaseManager:
    db = DatabaseManager(tmp_db_path)
    db.initialize()
    yield db
    db.close()


def _insert_case(db: DatabaseManager, case_id: str = "CS-2026-001", cx_1_val: str | None = None) -> None:
    db.connection.execute(
        "INSERT INTO cs_cases (case_id, subject, status, priority, replied,"
        " sent_time, created_at, updated_at)"
        " VALUES (?,?,?,?,?,?,?,?)",
        (
            case_id, "測試主旨", "處理中", "中", "否",
            "2026/03/26 10:00:00", "2026/03/26 10:00:00", "2026/03/26 10:00:00",
        ),
    )
    if cx_1_val is not None:
        db.connection.execute("UPDATE cs_cases SET cx_1=? WHERE case_id=?", (cx_1_val, case_id))
    db.connection.commit()


class TestTrackingTableWithCustomCols:
    def test_custom_col_header_in_tracking_table(self, db: DatabaseManager, tmp_path: Path):
        ccr = CustomColumnRepository(db.connection)
        ccr.add_column_to_cases("cx_1")
        ccr.insert("cx_1", "特殊備註", 1)
        _insert_case(db, cx_1_val="特殊值ABC")

        engine = ReportEngine(db.connection)
        out = tmp_path / "report.xlsx"
        engine.generate_tracking_table("2026/03/01", "2026/03/31", out)

        wb = openpyxl.load_workbook(out)
        ws2 = wb["問題追蹤總表"]
        headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
        assert "特殊備註" in headers

    def test_custom_col_value_in_tracking_table(self, db: DatabaseManager, tmp_path: Path):
        ccr = CustomColumnRepository(db.connection)
        ccr.add_column_to_cases("cx_1")
        ccr.insert("cx_1", "特殊備註", 1)
        _insert_case(db, cx_1_val="特殊值ABC")

        engine = ReportEngine(db.connection)
        out = tmp_path / "report.xlsx"
        engine.generate_tracking_table("2026/03/01", "2026/03/31", out)

        wb = openpyxl.load_workbook(out)
        ws2 = wb["問題追蹤總表"]
        headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
        col_idx = headers.index("特殊備註") + 1
        data_val = ws2.cell(row=2, column=col_idx).value
        assert data_val == "特殊值ABC"

    def test_company_sheet_includes_custom_col(self, db: DatabaseManager, tmp_path: Path):
        """個別公司頁籤也需包含自訂欄標題。"""
        db.connection.execute(
            "INSERT INTO companies (company_id, name) VALUES (?,?)",
            ("COMP-001", "測試公司"),
        )
        db.connection.commit()
        ccr = CustomColumnRepository(db.connection)
        ccr.add_column_to_cases("cx_1")
        ccr.insert("cx_1", "特殊備註", 1)
        _insert_case(db, cx_1_val="公司特殊值")
        db.connection.execute("UPDATE cs_cases SET company_id='COMP-001'")
        db.connection.commit()

        engine = ReportEngine(db.connection)
        out = tmp_path / "report.xlsx"
        engine.generate_tracking_table("2026/03/01", "2026/03/31", out)

        wb = openpyxl.load_workbook(out)
        company_sheets = [s for s in wb.sheetnames if "測試公司" in s or "COMP" in s]
        assert len(company_sheets) >= 1
        ws_c = wb[company_sheets[0]]
        headers = [ws_c.cell(row=2, column=c).value for c in range(1, ws_c.max_column + 1)]
        assert "特殊備註" in headers
