"""Tests for ReportEngine."""

from pathlib import Path

import openpyxl
import pytest

from hcp_cms.core.report_engine import ReportEngine
from hcp_cms.data.database import DatabaseManager
from hcp_cms.data.models import Case, Company, MantisTicket, QAKnowledge
from hcp_cms.data.repositories import CaseRepository, CompanyRepository, MantisRepository, QARepository


@pytest.fixture
def db(tmp_db_path: Path) -> DatabaseManager:
    db = DatabaseManager(tmp_db_path)
    db.initialize()
    yield db
    db.close()


@pytest.fixture
def seeded_db(db: DatabaseManager) -> DatabaseManager:
    comp_repo = CompanyRepository(db.connection)
    case_repo = CaseRepository(db.connection)
    qa_repo = QARepository(db.connection)

    comp_repo.insert(Company(company_id="C1", name="日月光", domain="ase.com"))
    comp_repo.insert(Company(company_id="C2", name="欣興", domain="uni.com"))

    case_repo.insert(Case(case_id="CS-2026-001", subject="薪資問題", company_id="C1",
                          status="處理中", priority="高", sent_time="2026/03/10 09:00",
                          system_product="HCP", issue_type="BUG"))
    case_repo.insert(Case(case_id="CS-2026-002", subject="請假設定", company_id="C1",
                          status="已回覆", sent_time="2026/03/15 10:00",
                          actual_reply="2026/03/15 14:00", system_product="HCP", issue_type="操作異常"))
    case_repo.insert(Case(case_id="CS-2026-003", subject="客製化需求", company_id="C2",
                          status="處理中", sent_time="2026/03/20 11:00",
                          issue_type="客制需求", system_product="HCP"))

    qa_repo.insert(QAKnowledge(qa_id="QA-202603-001", question="如何計算", answer="進入模組"))

    return db


class TestReportEngine:
    def test_generate_tracking_table(self, seeded_db, tmp_path):
        engine = ReportEngine(seeded_db.connection)
        path = engine.generate_tracking_table("2026/03/01", "2026/03/31", tmp_path / "tracking.xlsx")
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_tracking_table_sheets(self, seeded_db, tmp_path):
        engine = ReportEngine(seeded_db.connection)
        path = engine.generate_tracking_table("2026/03/01", "2026/03/31", tmp_path / "tracking.xlsx")
        wb = openpyxl.load_workbook(str(path))
        sheet_names = wb.sheetnames
        assert "📋 客戶索引" in sheet_names
        assert "問題追蹤總表" in sheet_names
        assert "QA知識庫" in sheet_names
        assert "Mantis提單追蹤" in sheet_names
        wb.close()

    def test_tracking_table_case_data(self, seeded_db, tmp_path):
        engine = ReportEngine(seeded_db.connection)
        path = engine.generate_tracking_table("2026/03/01", "2026/03/31", tmp_path / "tracking.xlsx")
        wb = openpyxl.load_workbook(str(path))
        ws = wb["問題追蹤總表"]
        # Header + 3 data rows
        assert ws.max_row == 4  # 1 header + 3 cases
        wb.close()

    def test_tracking_table_custom_sheet(self, seeded_db, tmp_path):
        engine = ReportEngine(seeded_db.connection)
        path = engine.generate_tracking_table("2026/03/01", "2026/03/31", tmp_path / "tracking.xlsx")
        wb = openpyxl.load_workbook(str(path))
        assert "客制需求" in wb.sheetnames
        ws = wb["客制需求"]
        assert ws.max_row == 2  # 1 header + 1 custom case
        wb.close()

    def test_generate_monthly_report(self, seeded_db, tmp_path):
        engine = ReportEngine(seeded_db.connection)
        path = engine.generate_monthly_report("2026/03/01", "2026/03/31", tmp_path / "report.xlsx")
        assert path.exists()

    def test_monthly_report_kpi(self, seeded_db, tmp_path):
        engine = ReportEngine(seeded_db.connection)
        path = engine.generate_monthly_report("2026/03/01", "2026/03/31", tmp_path / "report.xlsx")
        wb = openpyxl.load_workbook(str(path))
        ws = wb["📊 月報摘要"]
        # ReportWriter 寫法：row 1 = 標題列（header），row 2 = 欄位標題，row 3 = 案件總數
        assert ws.cell(row=3, column=2).value == 3
        # Row 4: 已回覆 = 1
        assert ws.cell(row=4, column=2).value == 1
        wb.close()

    def test_monthly_report_customer_analysis_sheet(self, seeded_db, tmp_path):
        engine = ReportEngine(seeded_db.connection)
        path = engine.generate_monthly_report("2026/03/01", "2026/03/31", tmp_path / "report.xlsx")
        wb = openpyxl.load_workbook(str(path))
        assert "🏢 客戶分析" in wb.sheetnames
        ws = wb["🏢 客戶分析"]
        # 1 header + 2 companies (日月光, 欣興)
        assert ws.max_row == 3
        wb.close()

    def test_report_with_no_data(self, db, tmp_path):
        engine = ReportEngine(db.connection)
        path = engine.generate_monthly_report("2026/01/01", "2026/01/31", tmp_path / "empty.xlsx")
        assert path.exists()
        wb = openpyxl.load_workbook(str(path))
        ws = wb["📊 月報摘要"]
        # ReportWriter 寫法：row 3 = 案件總數
        assert ws.cell(row=3, column=2).value == 0  # total = 0
        wb.close()

    def test_report_header_style(self, seeded_db, tmp_path):
        engine = ReportEngine(seeded_db.connection)
        path = engine.generate_tracking_table("2026/03/01", "2026/03/31", tmp_path / "styled.xlsx")
        wb = openpyxl.load_workbook(str(path))
        ws = wb["問題追蹤總表"]
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.color.rgb == "00FFFFFF"  # white text
        wb.close()

    def test_customer_index_has_link_text(self, seeded_db):
        """客戶索引的快速連結欄位應包含各公司連結文字。"""
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_tracking_table("2026/03/01", "2026/03/31")
        index_rows = data["📋 客戶索引"]
        # 日月光 in row 1 (index 1)，連結文字在第 6 欄（index 5）
        assert "日月光" in str(index_rows[1][5])

    def test_company_sheet_has_back_link_text(self, seeded_db):
        """各公司頁籤第一列應有返回客戶索引的文字。"""
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_tracking_table("2026/03/01", "2026/03/31")
        ase_key = [k for k in data if "日月光" in k][0]
        assert data[ase_key][0][0] == "↩ 返回客戶索引"

    # ── build_tracking_table 測試 ──────────────────────────────────────────

    def test_build_tracking_table_returns_dict(self, seeded_db):
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_tracking_table("2026/03/01", "2026/03/31")
        assert isinstance(data, dict)
        assert "📋 客戶索引" in data
        assert "問題追蹤總表" in data
        assert "QA知識庫" in data
        assert "Mantis提單追蹤" in data

    def test_build_tracking_table_sheet_structure(self, seeded_db):
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_tracking_table("2026/03/01", "2026/03/31")
        index_sheet = data["📋 客戶索引"]
        assert index_sheet[0] == ["#", "公司名稱", "Email 域名", "聯絡方式", "案件數", "快速連結"]
        tracking_sheet = data["問題追蹤總表"]
        assert len(tracking_sheet) == 4  # 1 header + 3 data rows

    def test_build_tracking_table_custom_sheet(self, seeded_db):
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_tracking_table("2026/03/01", "2026/03/31")
        assert "客制需求" in data
        assert len(data["客制需求"]) == 2  # 1 header + 1 custom case

    def test_build_tracking_table_company_sheets(self, seeded_db):
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_tracking_table("2026/03/01", "2026/03/31")
        ase_key = [k for k in data if "日月光" in k]
        assert len(ase_key) == 1
        ase_rows = data[ase_key[0]]
        assert ase_rows[0][0] == "↩ 返回客戶索引"
        assert ase_rows[1][0] == "案件編號"
        assert len(ase_rows) == 4  # link + header + 2 cases

    # ── build_monthly_report 測試 ──────────────────────────────────────────

    def test_build_monthly_report_returns_dict(self, seeded_db):
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_monthly_report("2026/03/01", "2026/03/31")
        assert isinstance(data, dict)
        assert "📊 月報摘要" in data
        assert "📋 案件明細" in data
        assert "🏢 客戶分析" in data

    def test_build_monthly_report_kpi(self, seeded_db):
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_monthly_report("2026/03/01", "2026/03/31")
        summary = data["📊 月報摘要"]
        # summary[0] = title row, summary[1] = header row, summary[2:] = KPI data rows
        total_row = summary[2]
        assert total_row[0] == "案件總數"
        assert total_row[1] == 3

    def test_build_monthly_report_case_detail(self, seeded_db):
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_monthly_report("2026/03/01", "2026/03/31")
        detail = data["📋 案件明細"]
        assert detail[0][0] == "案件編號"  # header
        assert len(detail) == 4  # 1 header + 3 cases

    def test_build_monthly_report_customer_analysis(self, seeded_db):
        engine = ReportEngine(seeded_db.connection)
        data = engine.build_monthly_report("2026/03/01", "2026/03/31")
        analysis = data["🏢 客戶分析"]
        assert analysis[0] == ["客戶", "已回覆", "處理中", "合計"]
        assert len(analysis) == 3  # 1 header + 2 companies

    def test_build_monthly_report_no_data(self, db):
        engine = ReportEngine(db.connection)
        data = engine.build_monthly_report("2026/01/01", "2026/01/31")
        summary = data["📊 月報摘要"]
        total_row = summary[2]
        assert total_row[1] == 0


@pytest.fixture
def mantis_seeded_db(db: DatabaseManager) -> DatabaseManager:
    repo = MantisRepository(db.connection)
    repo.upsert(MantisTicket(
        ticket_id="MT-0001", summary="系統當機", status="assigned",
        priority="urgent", handler="王小明", last_updated="2026/03/13 10:00:00",
    ))
    repo.upsert(MantisTicket(
        ticket_id="MT-0002", summary="薪資計算錯誤", status="assigned",
        priority="normal", handler="李大華", last_updated="2026/03/24 10:00:00",
    ))
    repo.upsert(MantisTicket(
        ticket_id="MT-0003", summary="登入頁跑版", status="resolved",
        priority="low", handler="張三", last_updated="2026/03/10 10:00:00",
    ))
    return db


class TestBuildMantisSheet:
    def test_returns_list_of_dicts(self, mantis_seeded_db):
        engine = ReportEngine(mantis_seeded_db.connection)
        rows = engine.build_mantis_sheet()
        assert isinstance(rows, list)
        assert len(rows) == 3

    def test_each_row_has_category(self, mantis_seeded_db):
        engine = ReportEngine(mantis_seeded_db.connection)
        rows = engine.build_mantis_sheet()
        for row in rows:
            assert "category" in row
            assert row["category"] in ("closed", "salary", "high", "normal")

    def test_sorting_high_before_closed(self, mantis_seeded_db):
        """high 優先度排在 closed 之前。"""
        engine = ReportEngine(mantis_seeded_db.connection)
        rows = engine.build_mantis_sheet()
        categories = [r["category"] for r in rows]
        assert categories.index("high") < categories.index("closed")


class TestGenerateMonthlyReportWithMantis:
    def test_monthly_report_has_mantis_sheet(self, mantis_seeded_db, tmp_path):
        """generate_monthly_report() 生成的 Excel 應包含 Mantis 追蹤工作表。"""
        engine = ReportEngine(mantis_seeded_db.connection)
        path = engine.generate_monthly_report(
            "2026/03/01", "2026/03/31", tmp_path / "monthly.xlsx"
        )
        wb = openpyxl.load_workbook(str(path))
        assert "📌 Mantis 追蹤" in wb.sheetnames

    def test_mantis_sheet_row_count(self, mantis_seeded_db, tmp_path):
        """Mantis 工作表的資料列數應等於 ticket 總數。"""
        engine = ReportEngine(mantis_seeded_db.connection)
        path = engine.generate_monthly_report(
            "2026/03/01", "2026/03/31", tmp_path / "monthly.xlsx"
        )
        wb = openpyxl.load_workbook(str(path))
        ws = wb["📌 Mantis 追蹤"]
        # 第 1 列是表頭，第 2 列起是資料
        data_rows = ws.max_row - 1
        assert data_rows == 3


class TestBuildMonthlyReportMantisStats:
    def test_mantis_stats_in_summary_sheet(self, mantis_seeded_db):
        """build_monthly_report() 的月報摘要應包含 Mantis 統計區段。"""
        engine = ReportEngine(mantis_seeded_db.connection)
        data = engine.build_monthly_report("2026/03/01", "2026/03/31")
        summary_rows = data["📊 月報摘要"]
        # 檢查是否含有「Mantis 追蹤統計」標題
        flat = [str(cell) for row in summary_rows for cell in row]
        assert any("Mantis" in cell for cell in flat)

    def test_mantis_stats_count_correct(self, mantis_seeded_db):
        """Mantis 統計合計應等於 ticket 總數（3 筆）。"""
        engine = ReportEngine(mantis_seeded_db.connection)
        data = engine.build_monthly_report("2026/03/01", "2026/03/31")
        summary_rows = data["📊 月報摘要"]
        # 找到「合計」列
        total_row = next((r for r in summary_rows if r and r[0] == "合計"), None)
        assert total_row is not None
        assert total_row[1] == 3


@pytest.fixture
def overdue_seeded_db(db: DatabaseManager) -> DatabaseManager:
    """種子資料：5 個處理中案件，sent_time 距今 3/5/7/10/35 天，對應 5 個超時級別。"""
    from datetime import datetime, timedelta

    comp_repo = CompanyRepository(db.connection)
    case_repo = CaseRepository(db.connection)

    comp_repo.insert(Company(company_id="C-J", name="J 公司", domain="j.com", cs_staff_id="STAFF-jill"))
    comp_repo.insert(Company(company_id="C-Y", name="Y 公司", domain="y.com", cs_staff_id="STAFF-YOGA"))

    now = datetime.now()
    samples = [
        ("CS-T0", 3),   # tier 0 (3-4 天)
        ("CS-T1", 5),   # tier 1 (5-6 天)
        ("CS-T2", 7),   # tier 2 (7-9 天)
        ("CS-T3", 10),  # tier 3 (10-29 天)
        ("CS-T4", 35),  # tier 4 (30+ 天)
    ]
    for case_id, days_ago in samples:
        ts = (now - timedelta(days=days_ago)).strftime("%Y/%m/%d %H:%M:%S")
        case_repo.insert(Case(
            case_id=case_id, subject=f"卡 {days_ago} 天",
            company_id="C-J", status="處理中", sent_time=ts, handler="jill",
            reply_count=1,
        ))
    # 加 1 個未指派 + 1 個 reply_count=2 不算 reply_count=1
    ts_old = (now - timedelta(days=15)).strftime("%Y/%m/%d %H:%M:%S")
    case_repo.insert(Case(
        case_id="CS-NOASSIGN", subject="未指派", company_id="C-Y", status="處理中",
        sent_time=ts_old, handler=None, reply_count=1,
    ))
    case_repo.insert(Case(
        case_id="CS-R2", subject="回覆 2 次", company_id="C-J", status="處理中",
        sent_time=ts_old, handler="jill", reply_count=2,
    ))
    return db


class TestBuildMonthlyReportOverdueTracking:
    """超時統計加入月報（B 方案：截至產報日算卡天數，丙：摘要 + 新 sheet 兩處呈現）。"""

    def _report_range(self) -> tuple[str, str]:
        """產生涵蓋全部 seed 案件的日期區間（今天 - 60 天 ~ 今天）。"""
        from datetime import datetime, timedelta
        now = datetime.now()
        start = (now - timedelta(days=60)).strftime("%Y/%m/%d")
        end = now.strftime("%Y/%m/%d")
        return start, end

    def test_overdue_section_in_summary_sheet(self, overdue_seeded_db):
        """月報摘要 sheet 應含「超時案件分布」標題。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        data = engine.build_monthly_report(start, end)
        summary = data["📊 月報摘要"]
        flat = [str(cell) for row in summary for cell in row]
        assert any("超時" in cell for cell in flat), \
            f"應含「超時」相關標題，實際：{flat}"

    def test_overdue_tracking_sheet_exists(self, overdue_seeded_db):
        """應新增「⏰ 案件追蹤」sheet。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        data = engine.build_monthly_report(start, end)
        assert "⏰ 案件追蹤" in data, f"應含「⏰ 案件追蹤」sheet，實際：{list(data.keys())}"

    def test_overdue_5_tier_distribution(self, overdue_seeded_db):
        """5 級超時分布應正確：3-4/5-6/7-9/10-29/30+ 各 1 筆。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        data = engine.build_monthly_report(start, end)
        tracking = data["⏰ 案件追蹤"]
        flat_str = "\n".join(str(cell) for row in tracking for cell in row if cell)
        # 5 級標籤都應出現
        assert "3-4 天" in flat_str
        assert "5-6 天" in flat_str
        assert "7-9 天" in flat_str
        assert "10-29 天" in flat_str
        assert "30+ 天" in flat_str

    def test_overdue_tracking_includes_unassigned_count(self, overdue_seeded_db):
        """追蹤 sheet 應含未指派案件統計。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        data = engine.build_monthly_report(start, end)
        tracking = data["⏰ 案件追蹤"]
        flat_str = "\n".join(str(cell) for row in tracking for cell in row if cell)
        assert "未指派" in flat_str

    def test_overdue_tracking_includes_reply_count_1(self, overdue_seeded_db):
        """追蹤 sheet 應含回覆 1 次的統計。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        data = engine.build_monthly_report(start, end)
        tracking = data["⏰ 案件追蹤"]
        flat_str = "\n".join(str(cell) for row in tracking for cell in row if cell)
        assert "回覆 1 次" in flat_str or "回覆次數" in flat_str

    def test_overdue_tracking_in_generated_xlsx(self, overdue_seeded_db, tmp_path):
        """generate_monthly_report 產出的 xlsx 應含「⏰ 案件追蹤」sheet。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        path = engine.generate_monthly_report(start, end, tmp_path / "monthly.xlsx")
        wb = openpyxl.load_workbook(str(path))
        assert "⏰ 案件追蹤" in wb.sheetnames
        wb.close()

    def test_per_handler_sheet_exists(self, overdue_seeded_db):
        """應新增「👥 各客服案件」sheet。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        data = engine.build_monthly_report(start, end)
        assert "👥 各客服案件" in data, f"應含「👥 各客服案件」sheet，實際：{list(data.keys())}"

    def test_per_handler_sheet_groups_by_handler(self, overdue_seeded_db):
        """各客服案件 sheet 應按 handler 分組（jill / YOGA / Rebecca / 未指派）。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        data = engine.build_monthly_report(start, end)
        per_handler = data["👥 各客服案件"]
        flat = "\n".join(str(cell) for row in per_handler for cell in row if cell)
        # seed 有 6 個 jill 案件，應該都列出
        assert "jill" in flat
        assert "CS-T0" in flat or "CS-T1" in flat  # 至少有 jill 名下案件

    def test_per_handler_sheet_sorts_reply1_and_oldest_first(self, db: DatabaseManager):
        """各客服案件 sheet 內，reply_count=1 且卡最久的案件應排在每個 handler 段最前面
        （代表「還沒實質回覆客戶」+ 拖最久 = 最緊急）。"""
        from datetime import datetime, timedelta

        case_repo = CaseRepository(db.connection)
        now = datetime.now()

        # 同 handler 內 4 個案件：
        # CS-R1-30D：reply_count=1、卡 30 天 ← 應第 1（最緊急）
        # CS-R1-5D ：reply_count=1、卡 5 天  ← 應第 2
        # CS-R2-30D：reply_count=2、卡 30 天 ← 應第 3
        # CS-R2-5D ：reply_count=2、卡 5 天  ← 應第 4
        for cid, rc, days in [
            ("CS-R1-5D", 1, 5),
            ("CS-R2-30D", 2, 30),
            ("CS-R1-30D", 1, 30),
            ("CS-R2-5D", 2, 5),
        ]:
            ts = (now - timedelta(days=days)).strftime("%Y/%m/%d %H:%M:%S")
            case_repo.insert(Case(
                case_id=cid, subject=f"{rc}-{days}", status="處理中",
                handler="jill", sent_time=ts, reply_count=rc,
            ))

        start = (now - timedelta(days=60)).strftime("%Y/%m/%d")
        end = now.strftime("%Y/%m/%d")
        engine = ReportEngine(db.connection)
        data = engine.build_monthly_report(start, end)
        per_handler = data["👥 各客服案件"]

        # 取出 jill section 內的案件編號順序
        case_ids: list[str] = []
        in_jill = False
        for row in per_handler:
            if row and any("jill" in str(cell) for cell in row):
                in_jill = True
                continue
            if in_jill and row and isinstance(row[0], str) and row[0].startswith("CS-R"):
                case_ids.append(row[0])

        # 驗證順序：reply_count=1 在前，組內按卡幾天 desc
        assert case_ids == ["CS-R1-30D", "CS-R1-5D", "CS-R2-30D", "CS-R2-5D"], \
            f"預期排序錯誤，實際：{case_ids}"

    def test_tracking_filtered_by_report_period(self, db: DatabaseManager):
        """追蹤統計應依報表日期區間過濾（月報 = 該月新建案件的快照，每月獨立）。

        例：報表 2026/05/01 ~ 2026/05/12，2025/01 建的舊案不應出現在 Section 2。
        """
        from datetime import datetime, timedelta
        case_repo = CaseRepository(db.connection)
        now = datetime.now()
        old_ts = (now - timedelta(days=200)).strftime("%Y/%m/%d %H:%M:%S")
        recent_ts = (now - timedelta(days=3)).strftime("%Y/%m/%d %H:%M:%S")

        case_repo.insert(Case(
            case_id="CS-OLD", subject="200 天舊案", status="處理中",
            handler="jill", sent_time=old_ts, reply_count=1,
        ))
        case_repo.insert(Case(
            case_id="CS-NEW", subject="新案", status="處理中",
            handler="jill", sent_time=recent_ts, reply_count=1,
        ))

        # 短日期區間（最近 30 天）— 舊案在區間外
        start = (now - timedelta(days=30)).strftime("%Y/%m/%d")
        end = now.strftime("%Y/%m/%d")
        engine = ReportEngine(db.connection)
        data = engine.build_monthly_report(start, end)

        tracking = data["⏰ 案件追蹤"]
        jill_row = next(
            (r for r in tracking if r and isinstance(r[0], str) and r[0] == "jill"),
            None,
        )
        assert jill_row is not None, "找不到 jill 列"
        # 「處理中」應為 1（只含本期新案，舊案排除）
        # 欄位：處理人, 處理中, 3-4, 5-6, 7-9, 10-29, 30+, 回覆 1 次, 最久卡天, 本期完成, 本期平均
        assert jill_row[1] == 1, f"處理中應為 1（舊案不計），實際：{jill_row[1]}"
        # 最久卡天應約為 3 天（新案）
        assert jill_row[8] < 30, f"最久卡天應 < 30（不含舊案），實際：{jill_row[8]}"

    def test_section2_has_split_efficiency_columns(self, db: DatabaseManager):
        """Section 2 應含「本期完成」「平均首回」「平均處理」3 欄（拆 FRT 與 total）。"""
        from datetime import datetime, timedelta

        from hcp_cms.data.repositories import CompanyRepository
        case_repo = CaseRepository(db.connection)
        comp_repo = CompanyRepository(db.connection)
        comp_repo.insert(Company(company_id="C-1", name="A", domain="a.com"))
        now = datetime.now()
        case_repo.insert(Case(
            case_id="CS-DONE", subject="完成案", status="已完成",
            handler="YOGA", company_id="C-1",
            sent_time=(now - timedelta(days=5)).strftime("%Y/%m/%d %H:%M:%S"),
            actual_reply=(now - timedelta(days=4, hours=12)).strftime("%Y/%m/%d %H:%M:%S"),
        ))

        start = (now - timedelta(days=30)).strftime("%Y/%m/%d")
        end = now.strftime("%Y/%m/%d")
        engine = ReportEngine(db.connection)
        data = engine.build_monthly_report(start, end)
        tracking = data["⏰ 案件追蹤"]
        header_row = next(
            (r for r in tracking if r and isinstance(r[0], str) and r[0] == "處理人"),
            None,
        )
        assert header_row is not None
        header_str = " ".join(str(c) for c in header_row)
        # 期望含「平均首回」和「平均處理」2 個欄位
        assert "平均首回" in header_str, f"應含「平均首回」，實際表頭：{header_str}"
        assert "平均處理" in header_str, f"應含「平均處理」，實際表頭：{header_str}"

    def test_section2_has_throughput_column(self, db: DatabaseManager):
        """Section 2 應含「本期完成」欄位。"""
        from datetime import datetime, timedelta

        from hcp_cms.data.repositories import CompanyRepository
        case_repo = CaseRepository(db.connection)
        comp_repo = CompanyRepository(db.connection)
        comp_repo.insert(Company(company_id="C-1", name="A", domain="a.com"))
        now = datetime.now()
        case_repo.insert(Case(
            case_id="CS-Y-DONE", subject="YOGA 完成案", status="已完成",
            handler="YOGA", company_id="C-1",
            sent_time=(now - timedelta(days=5)).strftime("%Y/%m/%d %H:%M:%S"),
            actual_reply=(now - timedelta(days=3)).strftime("%Y/%m/%d %H:%M:%S"),
        ))

        start = (now - timedelta(days=30)).strftime("%Y/%m/%d")
        end = now.strftime("%Y/%m/%d")
        engine = ReportEngine(db.connection)
        data = engine.build_monthly_report(start, end)
        tracking = data["⏰ 案件追蹤"]

        header_row = next(
            (r for r in tracking if r and isinstance(r[0], str) and r[0] == "處理人"),
            None,
        )
        assert header_row is not None
        header_str = " ".join(str(c) for c in header_row)
        assert "本期完成" in header_str

    def test_per_handler_divider_shows_5_tier_distribution(self, overdue_seeded_db):
        """各客服 section 標題應顯示 5 級超時分布 + 回覆 1 次 + 最久卡天。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        data = engine.build_monthly_report(start, end)
        per_handler = data["👥 各客服案件"]
        # 找 jill 的 section divider
        jill_divider = next(
            (str(r[0]) for r in per_handler if r and r[0] and "jill" in str(r[0]) and "━━" in str(r[0])),
            None,
        )
        assert jill_divider is not None, "找不到 jill 的 section divider"
        # 應包含 5 級標籤、回覆 1 次、最久卡
        for label in ["3-4 天", "5-6 天", "7-9 天", "10-29 天", "30+ 天", "回覆 1 次", "最久卡"]:
            assert label in jill_divider, f"divider 應含「{label}」，實際：{jill_divider}"

    def test_section2_handler_stats_5_tier_columns(self, overdue_seeded_db):
        """⏰ 案件追蹤 Section 2 工作量表應有 5 級欄位 + 回覆 1 次 + 最久卡天。"""
        engine = ReportEngine(overdue_seeded_db.connection)
        start, end = self._report_range()
        data = engine.build_monthly_report(start, end)
        tracking = data["⏰ 案件追蹤"]
        # 找「處理人 ...」表頭列
        header_row = next(
            (r for r in tracking if r and isinstance(r[0], str) and r[0] == "處理人"),
            None,
        )
        assert header_row is not None, "找不到處理人表頭列"
        header_str = " ".join(str(c) for c in header_row)
        for label in ["3-4 天", "5-6 天", "7-9 天", "10-29 天", "30+ 天", "回覆 1 次", "最久卡"]:
            assert label in header_str, f"Section 2 表頭應含「{label}」"

    def test_per_handler_sheet_excludes_system_companies_unassigned(self, db: DatabaseManager):
        """無 handler 但屬資通電腦/Mantis 的案件不應出現在「未指派」段。"""
        comp_repo = CompanyRepository(db.connection)
        comp_repo.insert(Company(company_id="C-ARES", name="資通電腦", domain="ares.com.tw"))
        comp_repo.insert(Company(company_id="C-MANTIS", name="Mantis", domain="mantis"))
        comp_repo.insert(Company(company_id="C-CUST", name="一般客戶", domain="cust.com"))

        case_repo = CaseRepository(db.connection)
        from datetime import datetime, timedelta
        ts = (datetime.now() - timedelta(days=5)).strftime("%Y/%m/%d %H:%M:%S")
        case_repo.insert(Case(
            case_id="CS-ARES-X", subject="資通內部", company_id="C-ARES",
            status="處理中", sent_time=ts, handler=None,
        ))
        case_repo.insert(Case(
            case_id="CS-MANTIS-X", subject="Mantis 通知", company_id="C-MANTIS",
            status="處理中", sent_time=ts, handler=None,
        ))
        case_repo.insert(Case(
            case_id="CS-CUST-X", subject="一般客戶未指派", company_id="C-CUST",
            status="處理中", sent_time=ts, handler=None,
        ))

        from datetime import datetime, timedelta
        now = datetime.now()
        start = (now - timedelta(days=60)).strftime("%Y/%m/%d")
        end = now.strftime("%Y/%m/%d")
        engine = ReportEngine(db.connection)
        data = engine.build_monthly_report(start, end)
        per_handler = data["👥 各客服案件"]
        flat = "\n".join(str(cell) for row in per_handler for cell in row if cell)
        assert "CS-ARES-X" not in flat, "資通電腦案件不應列入未指派"
        assert "CS-MANTIS-X" not in flat, "Mantis 案件不應列入未指派"
        assert "CS-CUST-X" in flat, "一般客戶未指派案件應列出"
