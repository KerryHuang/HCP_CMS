"""Tests for ReportWriter."""

from pathlib import Path

import openpyxl

from hcp_cms.core.report_writer import ReportWriter


class TestReportWriter:
    def test_write_excel_creates_file(self, tmp_path: Path):
        data = {"Sheet1": [["A", "B"], [1, 2]]}
        path = tmp_path / "test.xlsx"
        ReportWriter.write_excel(data, path)
        assert path.exists()

    def test_write_excel_sheet_names(self, tmp_path: Path):
        data = {
            "📋 客戶索引": [["#", "名稱"], [1, "A"]],
            "問題追蹤總表": [["案件編號"], ["CS-001"]],
        }
        path = tmp_path / "test.xlsx"
        ReportWriter.write_excel(data, path)
        wb = openpyxl.load_workbook(str(path))
        assert wb.sheetnames == ["📋 客戶索引", "問題追蹤總表"]
        wb.close()

    def test_write_excel_header_style(self, tmp_path: Path):
        data = {"Sheet1": [["A", "B"], [1, 2]]}
        path = tmp_path / "test.xlsx"
        ReportWriter.write_excel(data, path)
        wb = openpyxl.load_workbook(str(path))
        cell = wb["Sheet1"].cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.color.rgb == "00FFFFFF"
        wb.close()

    def test_write_excel_data_rows(self, tmp_path: Path):
        data = {"Sheet1": [["Name", "Value"], ["Alice", 10], ["Bob", 20]]}
        path = tmp_path / "test.xlsx"
        ReportWriter.write_excel(data, path)
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Sheet1"]
        assert ws.cell(row=2, column=1).value == "Alice"
        assert ws.cell(row=3, column=2).value == 20
        assert ws.max_row == 3
        wb.close()

    def test_write_excel_alternating_row_fill(self, tmp_path: Path):
        data = {"Sheet1": [["A"], ["r1"], ["r2"], ["r3"]]}
        path = tmp_path / "test.xlsx"
        ReportWriter.write_excel(data, path)
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Sheet1"]
        # Row 2 (even) should have alt fill
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00F9FAFB"
        # Row 3 (odd) should not have alt fill
        assert ws.cell(row=3, column=1).fill.start_color.rgb != "00F9FAFB"
        wb.close()

    def test_write_excel_empty_data(self, tmp_path: Path):
        data = {"EmptySheet": [["Header"]]}
        path = tmp_path / "test.xlsx"
        ReportWriter.write_excel(data, path)
        wb = openpyxl.load_workbook(str(path))
        assert wb["EmptySheet"].max_row == 1
        wb.close()


class TestAppendMantisSheet:
    def test_mantis_sheet_created(self, tmp_path):
        """append_mantis_sheet() 在既有 workbook 新增工作表。"""
        path = tmp_path / "report.xlsx"
        # 先建立一個基礎 workbook
        ReportWriter.write_excel({"摘要": [["欄位"], ["值"]]}, path)

        mantis_rows = [
            {"ticket_id": "MT-001", "summary": "問題A", "status": "assigned",
             "priority": "urgent", "unresolved_days": "5 天",
             "last_updated": "2026/03/26", "handler": "王小明", "category": "high"},
        ]
        ReportWriter.append_mantis_sheet(path, "📌 Mantis 追蹤", mantis_rows)

        wb = openpyxl.load_workbook(str(path))
        assert "📌 Mantis 追蹤" in wb.sheetnames

    def test_mantis_high_row_fill_color(self, tmp_path):
        """high 分類的列背景色應為 #450a0a。"""
        path = tmp_path / "report.xlsx"
        ReportWriter.write_excel({"摘要": [["欄位"], ["值"]]}, path)

        mantis_rows = [
            {"ticket_id": "MT-001", "summary": "急件", "status": "assigned",
             "priority": "urgent", "unresolved_days": "5 天",
             "last_updated": "2026/03/26", "handler": "王小明", "category": "high"},
        ]
        ReportWriter.append_mantis_sheet(path, "📌 Mantis 追蹤", mantis_rows)

        wb = openpyxl.load_workbook(str(path))
        ws = wb["📌 Mantis 追蹤"]
        # 第 1 列為表頭，第 2 列為資料
        fill = ws.cell(row=2, column=1).fill
        assert fill.fgColor.rgb.upper().endswith("450A0A")
